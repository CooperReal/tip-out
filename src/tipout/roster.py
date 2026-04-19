from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass
class Employee:
    canonical: str
    role: str


@dataclass
class Roster:
    employees: dict[str, Employee]
    aliases: dict[str, str]  # raw -> canonical

    def __post_init__(self):
        # Case-insensitive lookup indexes: map lowercase-stripped key -> canonical.
        self._employees_ci = {k.strip().lower(): k for k in self.employees}
        self._aliases_ci = {k.strip().lower(): v for k, v in self.aliases.items()}

    def resolve(self, raw: str) -> str | None:
        if not isinstance(raw, str):
            return None
        key = raw.strip().lower()
        if key in self._employees_ci:
            return self._employees_ci[key]
        return self._aliases_ci.get(key)


def load_roster(path: Path) -> Roster:
    wb = load_workbook(path, data_only=True)
    emps: dict[str, Employee] = {}
    for row in wb["Employees"].iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        name = row[0]
        role = row[1] if len(row) > 1 else None
        emps[name] = Employee(canonical=name, role=role or "")
    aliases: dict[str, str] = {}
    for row in wb["Name Aliases"].iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not row[1]:
            continue
        aliases[row[0]] = row[1]
    return Roster(employees=emps, aliases=aliases)
