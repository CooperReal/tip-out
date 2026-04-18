"""Anomaly checks + report writer.

Each check is a pure function over its inputs that returns a ``list[Anomaly]``
without raising. ``check_all`` aggregates every check in this module and
returns a deterministically-sorted list. ``write_report`` emits an
``anomaly_report.xlsx`` with one tab per pay period (append-only).
"""
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date as _date
from pathlib import Path

from openpyxl import Workbook, load_workbook

from tipout.hours import HoursEntry
from tipout.period import PayPeriod
from tipout.pos_parser import ShiftRow
from tipout.roster import Roster


@dataclass(frozen=True)
class Anomaly:
    """One flagged observation from an anomaly check."""

    kind: str
    severity: str
    date: _date | None
    employee: str | None
    message: str
    details: dict = field(default_factory=dict)


# ----------------------------------------------------------------------------
# 1. tip pool imbalance
# ----------------------------------------------------------------------------
def tip_pool_imbalance(shift_rows: list[ShiftRow]) -> list[Anomaly]:
    """Server-tier SA tip-out distributed should match support-tier net tips received.

    A "server-tier" row has ``cc_tips > 0``; "support-tier" has ``cc_tips == 0
    and net_tip > 0``. Mismatch > $0.50 on a given date is flagged.
    """
    by_date: dict[_date, list[ShiftRow]] = defaultdict(list)
    for r in shift_rows:
        by_date[r.date].append(r)

    out: list[Anomaly] = []
    for d in sorted(by_date.keys()):
        day_rows = by_date[d]
        sa_distributed = sum(r.sa_tip_out for r in day_rows if r.cc_tips > 0)
        support_received = sum(
            r.net_tip for r in day_rows if r.cc_tips == 0 and r.net_tip > 0
        )
        diff = sa_distributed - support_received
        if abs(diff) > 0.50:
            out.append(
                Anomaly(
                    kind="tip_pool_imbalance",
                    severity="warn",
                    date=d,
                    employee=None,
                    message=(
                        f"SA tip-out distributed (${sa_distributed:.2f}) does not "
                        f"match support-tier net tips received (${support_received:.2f}); "
                        f"diff ${diff:+.2f}"
                    ),
                    details={
                        "sa_distributed": round(sa_distributed, 2),
                        "support_received": round(support_received, 2),
                        "diff": round(diff, 2),
                    },
                )
            )
    return out


# ----------------------------------------------------------------------------
# 2. server math
# ----------------------------------------------------------------------------
def server_math(shift_rows: list[ShiftRow]) -> list[Anomaly]:
    """For server-tier rows, verify net_tip == cc_tips + party - total_tip_out (+/- $0.01).

    Skip rows flagged as party with zero total_tip_out (Patrick-party case).
    """
    out: list[Anomaly] = []
    for r in shift_rows:
        if r.cc_tips <= 0:
            continue
        if r.is_party and r.total_tip_out == 0:
            continue
        expected = r.cc_tips + r.party - r.total_tip_out
        diff = r.net_tip - expected
        if abs(diff) > 0.01:
            out.append(
                Anomaly(
                    kind="server_math",
                    severity="warn",
                    date=r.date,
                    employee=r.canonical_name,
                    message=(
                        f"net_tip ${r.net_tip:.2f} != cc_tips ${r.cc_tips:.2f} + "
                        f"party ${r.party:.2f} - total_tip_out ${r.total_tip_out:.2f} "
                        f"(expected ${expected:.2f}, diff ${diff:+.2f})"
                    ),
                    details={
                        "cc_tips": r.cc_tips,
                        "party": r.party,
                        "total_tip_out": r.total_tip_out,
                        "net_tip": r.net_tip,
                        "expected": round(expected, 2),
                        "diff": round(diff, 2),
                    },
                )
            )
    return out


# ----------------------------------------------------------------------------
# 3. outlier rate
# ----------------------------------------------------------------------------
def outlier_rate(
    shift_rows: list[ShiftRow],
    hours_entries: list[HoursEntry],
    period: PayPeriod,
) -> list[Anomaly]:
    """Per employee in ``period``, compute $/hr and flag if < $10 or > $100."""
    tips_by: dict[str, float] = defaultdict(float)
    hours_by: dict[str, float] = defaultdict(float)

    for r in shift_rows:
        if r.canonical_name and period.start <= r.date <= period.end:
            tips_by[r.canonical_name] += r.net_tip
    for h in hours_entries:
        if period.start <= h.date <= period.end:
            hours_by[h.canonical] += h.hours

    out: list[Anomaly] = []
    for emp in sorted(tips_by.keys() | hours_by.keys()):
        hrs = hours_by.get(emp, 0.0)
        tips = tips_by.get(emp, 0.0)
        if hrs <= 0:
            continue
        rate = tips / hrs
        if rate < 10.0 or rate > 100.0:
            out.append(
                Anomaly(
                    kind="outlier_rate",
                    severity="warn",
                    date=None,
                    employee=emp,
                    message=(
                        f"Effective rate ${rate:.2f}/hr (tips ${tips:.2f} / "
                        f"{hrs:.2f} hrs) outside expected $10-$100 band"
                    ),
                    details={
                        "tips": round(tips, 2),
                        "hours": round(hrs, 2),
                        "rate": round(rate, 2),
                    },
                )
            )
    return out


# ----------------------------------------------------------------------------
# 4. duplicate rows
# ----------------------------------------------------------------------------
def duplicate_rows(shift_rows: list[ShiftRow]) -> list[Anomaly]:
    """Same canonical_name + same date appearing 2+ times in the input."""
    groups: dict[tuple[str, _date], list[ShiftRow]] = defaultdict(list)
    for r in shift_rows:
        if not r.canonical_name:
            continue
        groups[(r.canonical_name, r.date)].append(r)

    out: list[Anomaly] = []
    for (canon, d), rs in sorted(groups.items()):
        if len(rs) < 2:
            continue
        nets = [round(r.net_tip, 2) for r in rs]
        out.append(
            Anomaly(
                kind="duplicate_rows",
                severity="warn",
                date=d,
                employee=canon,
                message=f"appears {len(rs)} times with net tips {nets}",
                details={"count": len(rs), "net_tips": nets},
            )
        )
    return out


# ----------------------------------------------------------------------------
# 5. period drop
# ----------------------------------------------------------------------------
def _tab_to_end_date(tab: str) -> _date | None:
    """Parse a summary tab name like '12.29 to 01.11.2026' to its end date."""
    try:
        left, right = tab.split(" to ")
        end_parts = right.split(".")
        if len(end_parts) != 3:
            return None
        m, d, y = int(end_parts[0]), int(end_parts[1]), int(end_parts[2])
        return _date(y, m, d)
    except (ValueError, AttributeError):
        return None


def period_drop(
    shift_rows: list[ShiftRow],
    roster: Roster,
    summary_path: Path,
    current_period: PayPeriod,
) -> list[Anomaly]:
    """Flag employees whose current period total dropped > 60% vs rolling avg of prior 4 periods."""
    _ = roster  # not currently used; reserved for active-period filtering
    if not summary_path.exists():
        return []

    # Aggregate current-period net tips per employee.
    current_totals: dict[str, float] = defaultdict(float)
    current_hours_seen: dict[str, bool] = defaultdict(bool)
    for r in shift_rows:
        if r.canonical_name and current_period.start <= r.date <= current_period.end:
            current_totals[r.canonical_name] += r.net_tip

    # Read prior tabs from summary file.
    try:
        wb = load_workbook(summary_path, read_only=True, data_only=True)
    except Exception:
        return []

    current_end = current_period.end
    prior_tabs: list[tuple[_date, str]] = []
    for tab in wb.sheetnames:
        end_d = _tab_to_end_date(tab)
        if end_d is None or end_d >= current_end:
            continue
        prior_tabs.append((end_d, tab))
    prior_tabs.sort(key=lambda t: t[0])
    last_n = prior_tabs[-4:]

    if not last_n:
        wb.close()
        return []

    # For each prior tab, extract canonical -> total from col A (1) and col 31.
    history: dict[str, list[float]] = defaultdict(list)
    for _end_d, tab in last_n:
        ws = wb[tab]
        for row in ws.iter_rows(min_row=5, values_only=True):
            if not row or not row[0]:
                continue
            canon = row[0]
            if not isinstance(canon, str):
                continue
            total = row[30] if len(row) > 30 else None
            if isinstance(total, (int, float)) and total > 0:
                history[canon].append(float(total))
    wb.close()

    # For per-employee hours-in-current-period gate, just check that their total > 0.
    out: list[Anomaly] = []
    for emp, current_total in sorted(current_totals.items()):
        prior = history.get(emp, [])
        if not prior:
            continue
        avg = sum(prior) / len(prior)
        if avg <= 0:
            continue
        if current_total < 0.40 * avg:
            drop_pct = (1.0 - current_total / avg) * 100.0
            out.append(
                Anomaly(
                    kind="period_drop",
                    severity="warn",
                    date=None,
                    employee=emp,
                    message=(
                        f"current period total ${current_total:.2f} is "
                        f"{drop_pct:.0f}% below rolling avg ${avg:.2f} "
                        f"(prior {len(prior)} periods)"
                    ),
                    details={
                        "current": round(current_total, 2),
                        "rolling_avg": round(avg, 2),
                        "prior_totals": [round(x, 2) for x in prior],
                        "drop_pct": round(drop_pct, 2),
                    },
                )
            )
    _ = current_hours_seen  # placeholder; could be wired to hours_entries if desired
    return out


# ----------------------------------------------------------------------------
# 6. write report
# ----------------------------------------------------------------------------
def _tab_name(period: PayPeriod) -> str:
    return (
        f"{period.start.month:02d}.{period.start.day:02d} to "
        f"{period.end.month:02d}.{period.end.day:02d}.{period.end.year}"
    )


def write_report(path: Path, anomalies: list[Anomaly], period: PayPeriod) -> None:
    """Emit anomaly_report.xlsx with one sheet per run. Append-only."""
    if path.exists():
        wb = load_workbook(path)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    tab = _tab_name(period)
    if tab in wb.sheetnames:
        raise ValueError(f"Anomaly report tab {tab!r} already exists")

    ws = wb.create_sheet(tab)
    ws.append(["Kind", "Severity", "Date", "Employee", "Message"])
    for a in anomalies:
        ws.append([a.kind, a.severity, a.date, a.employee, a.message])
    wb.save(path)


# ----------------------------------------------------------------------------
# 7. aggregator
# ----------------------------------------------------------------------------
def check_all(
    shift_rows: list[ShiftRow],
    hours_entries: list[HoursEntry],
    roster: Roster,
    summary_path: Path,
    period: PayPeriod,
) -> list[Anomaly]:
    """Run every check in this module; return a deterministically-sorted list."""
    anomalies: list[Anomaly] = []
    anomalies += tip_pool_imbalance(shift_rows)
    anomalies += server_math(shift_rows)
    anomalies += outlier_rate(shift_rows, hours_entries, period)
    anomalies += duplicate_rows(shift_rows)
    anomalies += period_drop(shift_rows, roster, summary_path, period)
    return sorted(
        anomalies,
        key=lambda a: (a.date or _date.min, a.kind, a.employee or ""),
    )
