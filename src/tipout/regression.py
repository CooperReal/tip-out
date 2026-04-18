"""Regression harness for tipout.

For each period fixture under tests/fixtures/periods/, runs the pipeline and
diffs actual output workbooks vs expected cell-by-cell.

Used by the `tipout test` CLI command and by pytest via test_regression.py.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
import shutil
import tempfile

from openpyxl import load_workbook

from tipout.config import Config
from tipout.period import PayPeriod
from tipout.runner import run


# Fixtures live under the repo's tests/fixtures/periods/. Resolve relative to
# this file: src/tipout/regression.py -> ../../tests/fixtures/periods.
FIXTURES_DIR = Path(__file__).resolve().parents[2] / "tests" / "fixtures" / "periods"


@dataclass(frozen=True)
class Diff:
    sheet: str
    cell: str          # e.g. "B5"
    expected: object
    actual: object
    message: str


def list_fixtures() -> list[Path]:
    """Return period-fixture directories (sorted)."""
    if not FIXTURES_DIR.exists():
        return []
    return sorted(p for p in FIXTURES_DIR.iterdir() if p.is_dir())


def run_fixture(fixture_dir: Path) -> list[Diff]:
    """Run the pipeline on fixture inputs and diff against expected/. Returns diffs; empty list = pass."""
    # Copy inputs to a scratch directory so the runner's output paths don't
    # pollute the fixture. Config.yaml in the fixture points to relative paths
    # that should resolve under the scratch dir.
    with tempfile.TemporaryDirectory() as tmp:
        scratch = Path(tmp)
        for name in (
            "input_pos.xlsx",
            "input_hours.xlsx",
            "input_roster.xlsx",
            "input_config.yaml",
        ):
            shutil.copy2(fixture_dir / name, scratch / name)

        config_path = scratch / "input_config.yaml"
        cfg = Config.load(config_path)

        # Derive period from fixture directory name, e.g. "2025-12-29_to_2026-01-11".
        # For synthetic/marker-named fixtures (e.g. "_synthetic") that don't encode
        # dates, fall back to a 14-day span starting at the config's anchor_date.
        period = _period_from_dirname(fixture_dir.name, fallback_anchor=cfg.anchor_date)

        # Run the pipeline
        run(cfg, scratch / "input_pos.xlsx", scratch / "input_hours.xlsx", period)

        # Diff produced summary vs expected
        diffs: list[Diff] = []
        diffs += _diff_workbook(
            cfg.summary_path,
            fixture_dir / "expected" / "summary.xlsx",
        )
        # Diff each per-employee file
        expected_per = fixture_dir / "expected" / "per_employee"
        if expected_per.exists():
            for exp_file in sorted(expected_per.glob("*.xlsx")):
                actual_file = cfg.per_employee_dir / exp_file.name
                diffs += _diff_workbook(actual_file, exp_file)
        return diffs


def _period_from_dirname(name: str, fallback_anchor: date | None = None) -> PayPeriod:
    """Parse a 14-day pay period from a fixture directory name.

    Real fixtures use the "<YYYY-MM-DD>_to_<YYYY-MM-DD>" naming convention.
    Synthetic/marker fixtures (e.g. "_synthetic") may not encode dates; for
    those, the caller can supply ``fallback_anchor`` (typically the config's
    anchor_date) and we use it as the period start.
    """
    bare = name.lstrip("_")
    start_str, sep, end_str = bare.partition("_to_")
    if sep:
        return PayPeriod.from_dates(
            date.fromisoformat(start_str), date.fromisoformat(end_str)
        )
    if fallback_anchor is not None:
        return PayPeriod.from_anchor(fallback_anchor, fallback_anchor)
    raise ValueError(
        f"Fixture directory name {name!r} must contain a "
        f"'<YYYY-MM-DD>_to_<YYYY-MM-DD>' segment."
    )


def _diff_workbook(actual_path: Path, expected_path: Path) -> list[Diff]:
    if not actual_path.exists():
        return [Diff("*", "*", "workbook exists", "missing", f"Expected file not produced: {actual_path}")]
    actual = load_workbook(actual_path, data_only=True)
    expected = load_workbook(expected_path, data_only=True)

    diffs: list[Diff] = []

    # Same sheet names?
    if set(actual.sheetnames) != set(expected.sheetnames):
        diffs.append(Diff(
            "*", "*",
            sorted(expected.sheetnames),
            sorted(actual.sheetnames),
            f"Sheet names differ in {actual_path.name}",
        ))

    for sheet_name in expected.sheetnames:
        if sheet_name not in actual.sheetnames:
            continue  # already reported above
        ws_a = actual[sheet_name]
        ws_e = expected[sheet_name]
        max_row = max(ws_a.max_row, ws_e.max_row)
        max_col = max(ws_a.max_column, ws_e.max_column)
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                va = ws_a.cell(row=r, column=c).value
                ve = ws_e.cell(row=r, column=c).value
                if not _cells_equal(va, ve):
                    diffs.append(Diff(
                        sheet=sheet_name,
                        cell=f"{ws_a.cell(row=r, column=c).coordinate}",
                        expected=ve,
                        actual=va,
                        message=f"Cell mismatch in {sheet_name}!{ws_a.cell(row=r, column=c).coordinate}",
                    ))
    return diffs


def _cells_equal(a, b, tol: float = 0.005) -> bool:
    """Compare two cell values with tolerance for floats and type coercion for date/datetime."""
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    # Date vs datetime coercion (openpyxl round-trips date as datetime)
    def _norm(v):
        if isinstance(v, datetime):
            return v.date() if v.time() == datetime.min.time() else v
        return v
    a2, b2 = _norm(a), _norm(b)
    if isinstance(a2, float) and isinstance(b2, float):
        return abs(a2 - b2) <= tol
    if isinstance(a2, (int, float)) and isinstance(b2, (int, float)):
        return abs(float(a2) - float(b2)) <= tol
    return a2 == b2
