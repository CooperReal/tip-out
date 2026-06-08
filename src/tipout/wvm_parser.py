"""Parser for the Watersound (WVM) day-per-tab daily worksheet.

Emits the same ``ShiftRow`` dataclass as the Surfing Deer parser so the downstream
runner/summary consume both restaurants identically. For the WVM summary only the
combined ``net_tip`` (col L) is needed; the AM/PM-split fields are intentionally left
at 0.0 (they are wrong wherever a split header is corrupted — see the design spec).
"""

from __future__ import annotations

import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Iterator

from openpyxl import load_workbook

from tipout.pos_parser import SchemaError, ShiftRow

# Header label (normalized: stripped + lowercased) -> canonical key.
_HEADER_ALIASES = {
    "am cc tips": "AM CC Tips",
    "pm cc tips": "PM CC Tips",
    "am staff tip out": "AM Staff Tip Out",
    "pm staff tip out": "PM Staff Tip Out",
    "am bar tipout": "AM Bar Tipout",
    "pm bar tipout": "PM Bar Tipout",
    "totaltip out": "TotalTip Out",
    "serv as": "Serv As",
    "bartender": "Bartender",
    "net tip": "Net tip",
}

# Real role-group labels (col A). Anything else is junk and must not become a role.
KNOWN_GROUPS = {"wait am", "bartndr", "host", "sa", "f runner", "to go"}

_TAB_DATE_RE = re.compile(r"^\s*(\d{1,2})\.(\d{1,2})\.(\d{2}|\d{4})\s*$")


class WvmFormatError(RuntimeError):
    """Raised when a workbook has no WVM day-tabs (likely the wrong file/restaurant)."""


def _parse_tab_date(sheet_name: str) -> date | None:
    """Parse a WVM tab name like '12.29.25' or '01.01.2026' (trailing spaces ok)."""
    m = _TAB_DATE_RE.match(sheet_name)
    if not m:
        return None
    mm, dd, yy = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if yy < 100:
        yy += 2000
    try:
        return date(yy, mm, dd)
    except ValueError:
        return None


def _find_header(ws) -> tuple[int, dict[str, int]]:
    """Locate the header row by content; return (row, {canonical_key: column})."""
    for r in range(1, min(ws.max_row, 15) + 1):
        found: dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                key = _HEADER_ALIASES.get(v.strip().lower())
                if key:
                    found.setdefault(key, c)
        if "Net tip" in found:
            return r, found
    raise SchemaError(f"No 'Net tip' header found in WVM sheet {ws.title!r}")


def _resolve_date(ws, tab_date: date, warn) -> date:
    """Tab name is authoritative; B3 is a cross-check. Warn on bad/disagreeing B3."""
    b3 = ws.cell(row=3, column=2).value
    if isinstance(b3, datetime):
        b3d = b3.date()
    elif isinstance(b3, date):
        b3d = b3
    else:
        b3d = None
    if b3d is None and b3 is not None:
        warn(f"{ws.title!r}: B3 {b3!r} is not a date; using tab-name date {tab_date}")
    elif b3d is not None and b3d != tab_date:
        warn(f"{ws.title!r}: B3 date {b3d} != tab-name date {tab_date}; using tab name")
    return tab_date


def _name_col(headers: dict[str, int]) -> int:
    cc = headers.get("AM CC Tips") or headers.get("PM CC Tips")
    return (cc - 1) if cc else 2  # name col is immediately left of CC; layout puts it at B


def _iter_rows(ws, header_row: int, name_col: int, net_col: int):
    """Yield (name, net, group) for each worker row; stop at the 'P/O' block.

    `group` carries down from the last KNOWN_GROUPS label (junk labels are ignored).
    """
    current_group = ""
    for r in range(header_row + 1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if isinstance(a, str):
            astr = a.strip()
            if astr.lower() == "p/o":
                return
            if astr.lower() in KNOWN_GROUPS:
                current_group = astr
        name = ws.cell(row=r, column=name_col).value
        if not isinstance(name, str) or not name.strip():
            continue
        net = ws.cell(row=r, column=net_col).value
        net = float(net) if isinstance(net, (int, float)) else 0.0
        yield name.strip(), net, current_group


def _day_tabs(wb):
    """Yield (sheet, tab_date) for every WVM day-tab (date-shaped name)."""
    for name in wb.sheetnames:
        d = _parse_tab_date(name)
        if d is not None:
            yield wb[name], d


def parse_workbook(path: Path) -> list[ShiftRow]:
    wb = load_workbook(path, data_only=True)
    rows: list[ShiftRow] = []
    warnings: list[str] = []
    n_tabs = 0
    for ws, tab_date in _day_tabs(wb):
        header_row, headers = _find_header(ws)
        n_tabs += 1
        the_date = _resolve_date(ws, tab_date, warnings.append)
        net_col = headers["Net tip"]
        name_col = _name_col(headers)
        for name, net, _group in _iter_rows(ws, header_row, name_col, net_col):
            if net == 0.0:
                continue
            rows.append(ShiftRow(
                date=the_date, raw_name=name,
                cc_tips=0.0, party=0.0, sa_tip_out=0.0, bar_tipout=0.0,
                total_tip_out=0.0, barback=0.0, bartender=0.0,
                net_tip=net, is_party=False,
            ))
    if n_tabs == 0:
        raise WvmFormatError(
            f"{path}: no WVM day-tabs found — is this a WVM daily workbook?"
        )
    for w in warnings:
        print(f"WARNING: {w}", file=sys.stderr)
    return rows


def read_day_net_totals(path: Path) -> dict[date, float]:
    """Per day-tab, the sheet's own total Net Tip — the totals row directly above the
    'P/O' block (the 'L56' figure). Used as an independent integrity cross-check.
    """
    wb = load_workbook(path, data_only=True)
    out: dict[date, float] = {}
    for ws, tab_date in _day_tabs(wb):
        header_row, headers = _find_header(ws)
        net_col = headers["Net tip"]
        po_row = None
        for r in range(header_row + 1, ws.max_row + 1):
            a = ws.cell(row=r, column=1).value
            if isinstance(a, str) and a.strip().lower() == "p/o":
                po_row = r
                break
        if po_row is None:
            continue
        v = ws.cell(row=po_row - 1, column=net_col).value
        if isinstance(v, (int, float)):
            out[tab_date] = float(v)
    return out


def iter_daily_names(path: Path) -> Iterator[tuple[str, str]]:
    """Yield (raw_name, group) for every worker row across all day-tabs.

    Does NOT skip zero-net rows (a person may have tips on another day). `group` is a
    KNOWN_GROUPS label or '' (junk/blank). Used by the roster harvester.
    """
    wb = load_workbook(path, data_only=True)
    for ws, _tab_date in _day_tabs(wb):
        header_row, headers = _find_header(ws)
        name_col = _name_col(headers)
        net_col = headers["Net tip"]
        for name, _net, group in _iter_rows(ws, header_row, name_col, net_col):
            yield name, group
