"""Per-employee tip-out workbooks.

One file per canonical employee under `<output_dir>/per-employee/<Name>.xlsx`.
One tab per pay period, appended (existing tabs untouched). Layout matches
the hand-done Yvonne.xlsx reference: Hours Worked at col B, tip columns
C..I, $/hr at col J on the totals row only.
"""

from collections import defaultdict
from datetime import date, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from tipout.period import PayPeriod
from tipout.pos_parser import ShiftRow
from tipout.summary import _tab_name

TIP_FORMAT = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'
DATE_FORMAT = "ddd m/d"
HOURS_FORMAT = "0.00"  # plain two-decimal number — hours are not currency

# (header label, ShiftRow attribute). Date column is special-cased.
TIP_COLUMNS: list[tuple[str, str]] = [
    ("CC Tips", "cc_tips"),
    ("SA Tip Out", "sa_tip_out"),
    ("Bar Tipout", "bar_tipout"),
    ("Total Tip Out", "total_tip_out"),
    ("Barback", "barback"),
    ("Bartender", "bartender"),
    ("Net Tip", "net_tip"),
]
HOURS_HEADER = "Hours Worked"
PER_HR_HEADER = "$/hr"
TOTAL_COLS = 1 + 1 + len(TIP_COLUMNS) + 1  # Date + Hours + 7 tips + $/hr = 10

# Layout (1-indexed Excel rows/cols).
TITLE_ROW = 1
HEADER_ROW = 2
FIRST_DATA_ROW = 3
DATE_COL = 1
HOURS_COL = 2
FIRST_TIP_COL = 3
NET_TIP_COL = 2 + len(TIP_COLUMNS)        # = 9 (col I)
PER_HR_COL = TOTAL_COLS                    # = 10 (col J)

COLUMN_WIDTHS: dict[str, float] = {
    "A": 11.0,   # Date
    "B": 12.0,   # Hours Worked
    "C": 11.0, "D": 11.0, "E": 11.0,
    "F": 13.0,   # Total Tip Out
    "G": 11.0, "H": 11.0,
    "I": 12.0,   # Net Tip
    "J": 9.0,    # $/hr
}

THIN = Side(style="thin", color="B0B0B0")
MEDIUM = Side(style="medium", color="000000")
HEADER_FILL = PatternFill("solid", fgColor="F2F2F2")


def _round2(x: float) -> float:
    return round(x, 2)


def _safe_filename(canonical: str) -> str:
    """Replace characters illegal in Windows filenames with '_'."""
    bad = '<>:"/\\|?*'
    return "".join("_" if c in bad else c for c in canonical).strip()


def build_grid(
    period: PayPeriod,
    canonical: str,
    shift_rows: list[ShiftRow],
    hours_by_date: dict[date, float] | None = None,
) -> list[list[Any]]:
    """Return a 2D grid for one employee's per-period tab.

    Layout: row 1 title, row 2 column headers, rows 3..16 = 14 day rows,
    row 17 = totals. Day rows always present (blanks where no shift).
    When ``hours_by_date`` is provided, col B is populated per-day and col J
    on the totals row holds Net Tip total ÷ Hours total (blank if Hours = 0).
    """
    hours_by_date = hours_by_date or {}

    in_period = [
        r
        for r in shift_rows
        if r.canonical_name == canonical
        and period.start <= r.date <= period.end
    ]

    # Sum by date — same-day duplicate rows are rare but safe to fold.
    by_date: dict[date, dict[str, float]] = defaultdict(
        lambda: {attr: 0.0 for _, attr in TIP_COLUMNS}
    )
    for r in in_period:
        for _, attr in TIP_COLUMNS:
            by_date[r.date][attr] += getattr(r, attr)

    dates_in_period = [period.start + timedelta(days=i) for i in range(14)]

    title = (
        f"{canonical}   Pay Period "
        f"{period.start.month:02d}.{period.start.day:02d} to "
        f"{period.end.month:02d}.{period.end.day:02d}.{period.end.year}"
    )

    row1: list[Any] = [None] * TOTAL_COLS
    row1[0] = title

    row2: list[Any] = (
        ["Date", HOURS_HEADER]
        + [label for label, _ in TIP_COLUMNS]
        + [PER_HR_HEADER]
    )

    grid: list[list[Any]] = [row1, row2]

    totals: dict[str, float] = {attr: 0.0 for _, attr in TIP_COLUMNS}
    hours_total = 0.0
    for d in dates_in_period:
        row: list[Any] = [None] * TOTAL_COLS
        row[0] = d
        hours = hours_by_date.get(d, 0.0)
        if hours:
            row[HOURS_COL - 1] = _round2(hours)
            hours_total += hours
        day_data = by_date.get(d)
        if day_data:
            for col_idx, (_, attr) in enumerate(TIP_COLUMNS, start=FIRST_TIP_COL - 1):
                v = _round2(day_data[attr])
                if v != 0:
                    row[col_idx] = v
                    totals[attr] += v
        # $/hr column (J) is always blank on day rows.
        grid.append(row)

    totals_row: list[Any] = [None] * TOTAL_COLS
    totals_row[0] = "Total"
    if hours_total:
        totals_row[HOURS_COL - 1] = _round2(hours_total)
    for col_idx, (_, attr) in enumerate(TIP_COLUMNS, start=FIRST_TIP_COL - 1):
        totals_row[col_idx] = _round2(totals[attr])
    net_tip_total = totals[TIP_COLUMNS[-1][1]]
    if hours_total > 0:
        totals_row[PER_HR_COL - 1] = round(net_tip_total / hours_total, 2)
    grid.append(totals_row)

    return grid


def _apply_styling(ws, totals_row: int) -> None:
    """Column widths, freeze panes, formats, borders, bold totals."""
    for letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[letter].width = width

    ws.freeze_panes = "A3"

    # Title
    ws.cell(row=TITLE_ROW, column=1).font = Font(name="Calibri", size=14, bold=True)
    ws.merge_cells(
        start_row=TITLE_ROW, start_column=1,
        end_row=TITLE_ROW, end_column=TOTAL_COLS,
    )

    # Header row: bold, gray fill, centered.
    header_font = Font(bold=True)
    centered = Alignment(horizontal="center")
    for c in range(1, TOTAL_COLS + 1):
        cell = ws.cell(row=HEADER_ROW, column=c)
        cell.font = header_font
        cell.fill = HEADER_FILL
        cell.alignment = centered

    # Date format on the Date column body cells.
    for r in range(FIRST_DATA_ROW, totals_row + 1):
        ws.cell(row=r, column=DATE_COL).number_format = DATE_FORMAT

    # Hours Worked (col B) is a plain number, not currency — day rows and total.
    for r in range(FIRST_DATA_ROW, totals_row + 1):
        ws.cell(row=r, column=HOURS_COL).number_format = HOURS_FORMAT

    # Currency format on the tip columns (C..I) and the $/hr cell on totals (J).
    for r in range(FIRST_DATA_ROW, totals_row + 1):
        for c in range(FIRST_TIP_COL, NET_TIP_COL + 1):
            ws.cell(row=r, column=c).number_format = TIP_FORMAT
    ws.cell(row=totals_row, column=PER_HR_COL).number_format = TIP_FORMAT

    # Bold totals row.
    bold = Font(bold=True)
    for c in range(1, TOTAL_COLS + 1):
        ws.cell(row=totals_row, column=c).font = bold

    # Borders on the entire data area.
    for r in range(HEADER_ROW, totals_row + 1):
        for c in range(1, TOTAL_COLS + 1):
            top = THIN
            bottom = THIN
            if r == totals_row:
                top = MEDIUM
            if r == totals_row - 1:
                bottom = MEDIUM
            ws.cell(row=r, column=c).border = Border(
                left=THIN, right=THIN, top=top, bottom=bottom,
            )


def append_period_tab_for_employee(
    output_dir: Path,
    period: PayPeriod,
    canonical: str,
    shift_rows: list[ShiftRow],
    hours_by_date: dict[date, float] | None = None,
) -> Path:
    """Append a new period tab to <output_dir>/per-employee/<canonical>.xlsx.

    Returns the file path written. Raises ValueError if the period tab already
    exists in that file.
    """
    per_emp_dir = output_dir / "per-employee"
    per_emp_dir.mkdir(parents=True, exist_ok=True)
    file_path = per_emp_dir / f"{_safe_filename(canonical)}.xlsx"

    if file_path.exists():
        wb = load_workbook(file_path)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    tab = _tab_name(period)
    if tab in wb.sheetnames:
        raise ValueError(
            f"Tab {tab!r} already exists in {file_path.name} — delete to re-run"
        )

    ws = wb.create_sheet(tab)
    grid = build_grid(period, canonical, shift_rows, hours_by_date=hours_by_date)
    for r, row_values in enumerate(grid, start=1):
        for c, val in enumerate(row_values, start=1):
            if val is not None:
                ws.cell(row=r, column=c, value=val)

    totals_row = HEADER_ROW + 14 + 1  # = 17
    _apply_styling(ws, totals_row)

    wb.save(file_path)
    return file_path
