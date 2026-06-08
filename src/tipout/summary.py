from collections import defaultdict
from datetime import date, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from tipout.period import PayPeriod
from tipout.pos_parser import ShiftRow
from tipout.roster import Roster

TIP_FORMAT = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'
DATE_FORMAT = "ddd m/d"  # e.g. "Mon 12/29" — easier to scan than mm-dd-yy

# Layout: col A = name, cols B..O = days 1..14, col P = total. No spacer columns.
COLUMN_WIDTHS: dict[str, float] = {
    "A": 26.0,
    "B": 10.5, "C": 10.5, "D": 10.5, "E": 10.5, "F": 10.5, "G": 10.5, "H": 10.5,
    "I": 10.5, "J": 10.5, "K": 10.5, "L": 10.5, "M": 10.5, "N": 10.5, "O": 10.5,
    "P": 12.0,
}

NAME_COL = 1            # A
FIRST_DAY_COL = 2       # B (day 1); day N at FIRST_DAY_COL + (N - 1)
TOTAL_COL = 16          # P
WEEK_DIVIDER_BEFORE_COL = FIRST_DAY_COL + 7  # I — left border of day 8 = week divider

TITLE_ROW = 1
DATE_ROW = 2            # was 3 in old layout; tightened
FIRST_EMPLOYEE_ROW = 3  # was 5

THIN = Side(style="thin", color="B0B0B0")
MEDIUM = Side(style="medium", color="000000")
HEADER_FILL = PatternFill("solid", fgColor="F2F2F2")


def _round2(x: float) -> float:
    return round(x, 2)


def build_grid(
    period: PayPeriod,
    shift_rows: list[ShiftRow],
    roster: Roster,
    restaurant_name: str = "Surfing Deer",
) -> list[list[Any]]:
    """Return a 2D grid of cell values for the 2-week summary tab.

    Layout: row 1 title, row 2 date header, rows 3..N employee rows (one per
    canonical with at least one in-period shift, in roster insertion order),
    row N+1 totals row. Columns: A name, B..O day 1..14, P grand total.
    """
    in_period = [
        r
        for r in shift_rows
        if period.start <= r.date <= period.end and r.canonical_name
    ]

    by_emp_date: dict[str, dict[date, float]] = defaultdict(dict)
    for r in in_period:
        by_emp_date[r.canonical_name][r.date] = (
            by_emp_date[r.canonical_name].get(r.date, 0.0) + r.net_tip
        )

    dates_in_period = [period.start + timedelta(days=i) for i in range(14)]
    title = (
        f"{restaurant_name} Tip outs "
        f"{period.start.month:02d}.{period.start.day:02d} to "
        f"{period.end.month:02d}.{period.end.day:02d}.{period.end.year}"
    )

    width = TOTAL_COL  # 16 columns total

    row1: list[Any] = [None] * width
    row1[0] = title

    row2: list[Any] = [None] * width
    row2[0] = "Employee"
    for i, d in enumerate(dates_in_period):
        row2[FIRST_DAY_COL - 1 + i] = d
    row2[TOTAL_COL - 1] = "Total"

    grid: list[list[Any]] = [row1, row2]

    daily_totals: dict[date, float] = {d: 0.0 for d in dates_in_period}

    for canon in roster.employees.keys():
        day_tips = by_emp_date.get(canon)
        if not day_tips:
            continue
        total = _round2(sum(day_tips.values()))

        row: list[Any] = [None] * width
        row[0] = canon
        for i, d in enumerate(dates_in_period):
            v = day_tips.get(d)
            if v is not None:
                rounded = _round2(v)
                row[FIRST_DAY_COL - 1 + i] = rounded
                daily_totals[d] += rounded
        row[TOTAL_COL - 1] = total
        grid.append(row)

    totals_row: list[Any] = [None] * width
    totals_row[0] = "Daily Total"
    grand = 0.0
    for i, d in enumerate(dates_in_period):
        col_total = _round2(daily_totals[d])
        totals_row[FIRST_DAY_COL - 1 + i] = col_total
        grand += col_total
    totals_row[TOTAL_COL - 1] = _round2(grand)
    grid.append(totals_row)

    return grid


def _tab_name(period: PayPeriod) -> str:
    return (
        f"{period.start.month:02d}.{period.start.day:02d} to "
        f"{period.end.month:02d}.{period.end.day:02d}.{period.end.year}"
    )


def _border_for(row: int, col: int, *, last_row: int, totals_row: int) -> Border:
    """Compose the per-cell border for a data cell.

    All data cells get thin gray on every side. The week divider (between day 7
    and day 8) is medium black. The totals row gets medium black above it.
    """
    left = THIN
    right = THIN
    top = THIN
    bottom = THIN

    if col == WEEK_DIVIDER_BEFORE_COL:
        left = MEDIUM
    if col == WEEK_DIVIDER_BEFORE_COL - 1:
        right = MEDIUM

    if row == totals_row:
        top = MEDIUM
    if row == totals_row - 1:
        bottom = MEDIUM

    return Border(left=left, right=right, top=top, bottom=bottom)


def _apply_styling(ws, num_employee_rows: int) -> None:
    """Apply column widths, freeze panes, formats, borders, and header styling."""
    for letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[letter].width = width

    ws.freeze_panes = "B3"  # lock col A and the title+header rows

    # Title row.
    title_cell = ws.cell(row=TITLE_ROW, column=NAME_COL)
    title_cell.font = Font(name="Calibri", size=14, bold=True)
    ws.merge_cells(
        start_row=TITLE_ROW, start_column=NAME_COL,
        end_row=TITLE_ROW, end_column=TOTAL_COL,
    )

    totals_row = FIRST_EMPLOYEE_ROW + num_employee_rows
    last_row = totals_row

    # Header row formatting + date format.
    header_font = Font(bold=True)
    centered = Alignment(horizontal="center")
    for c in range(NAME_COL, TOTAL_COL + 1):
        cell = ws.cell(row=DATE_ROW, column=c)
        cell.font = header_font
        cell.fill = HEADER_FILL
        if FIRST_DAY_COL <= c <= FIRST_DAY_COL + 13:
            cell.number_format = DATE_FORMAT
            cell.alignment = centered

    # Tip number format on employee + totals rows.
    for r in range(FIRST_EMPLOYEE_ROW, totals_row + 1):
        for c in range(FIRST_DAY_COL, TOTAL_COL + 1):
            ws.cell(row=r, column=c).number_format = TIP_FORMAT

    # Bold the totals row label + values.
    bold = Font(bold=True)
    for c in range(NAME_COL, TOTAL_COL + 1):
        ws.cell(row=totals_row, column=c).font = bold

    # Borders around the entire data area (header row + body + totals).
    for r in range(DATE_ROW, last_row + 1):
        for c in range(NAME_COL, TOTAL_COL + 1):
            ws.cell(row=r, column=c).border = _border_for(
                r, c, last_row=last_row, totals_row=totals_row
            )


def append_period_tab(
    summary_path: Path,
    period: PayPeriod,
    shift_rows: list[ShiftRow],
    roster: Roster,
    restaurant_name: str = "Surfing Deer",
) -> None:
    """Append a new period tab to the 2-week summary workbook, creating it if absent.

    Raises ValueError if the tab for this period already exists.
    """
    if summary_path.exists():
        wb = load_workbook(summary_path)
    else:
        summary_path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    tab_name = _tab_name(period)
    if tab_name in wb.sheetnames:
        raise ValueError(f"Tab {tab_name!r} already exists — delete to re-run")

    ws = wb.create_sheet(tab_name)
    grid = build_grid(period, shift_rows, roster, restaurant_name=restaurant_name)
    for r, row_values in enumerate(grid, start=1):
        for c, val in enumerate(row_values, start=1):
            if val is not None:
                ws.cell(row=r, column=c, value=val)

    # rows: 1 title + 1 header + N employees + 1 totals = N + 3
    num_employee_rows = len(grid) - 3
    _apply_styling(ws, num_employee_rows)

    wb.save(summary_path)
