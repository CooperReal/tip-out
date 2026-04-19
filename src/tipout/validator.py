"""Roster integrity checker.

Reads `roster.xlsx` and reports structural + semantic issues so the operator
can catch mistakes before they reach a `tipout run`.
"""

from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass(frozen=True)
class Issue:
    severity: str  # "error" | "warning"
    message: str


REQUIRED_EMPLOYEES_HEADER = "Canonical Name"
REQUIRED_ALIAS_HEADERS = ("Raw Name", "Canonical Name")


def validate_roster(path: Path) -> list[Issue]:
    """Return a list of issues found in the roster. Empty list = clean."""
    if not path.exists():
        return [Issue("error", f"Roster file not found: {path}")]

    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        return [Issue("error", f"Could not open roster: {type(e).__name__}: {e}")]

    issues: list[Issue] = []

    # --- Employees sheet ---
    if "Employees" not in wb.sheetnames:
        issues.append(Issue("error", "Missing required sheet: 'Employees'"))
        return issues

    emp_ws = wb["Employees"]
    header_a = emp_ws.cell(row=1, column=1).value
    if header_a != REQUIRED_EMPLOYEES_HEADER:
        issues.append(Issue(
            "error",
            f"Employees sheet: cell A1 must be {REQUIRED_EMPLOYEES_HEADER!r}, got {header_a!r}",
        ))

    canonicals: list[str] = []
    for r, row in enumerate(emp_ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue
        name = row[0]
        if not isinstance(name, str) or not name.strip():
            issues.append(Issue("error", f"Employees row {r}: blank or non-string canonical name"))
            continue
        canonicals.append(name.strip())

    dup_canonicals = [n for n, count in Counter(canonicals).items() if count > 1]
    for dup in dup_canonicals:
        issues.append(Issue("warning", f"Duplicate canonical in Employees: {dup!r}"))

    # --- Name Aliases sheet ---
    if "Name Aliases" not in wb.sheetnames:
        issues.append(Issue("error", "Missing required sheet: 'Name Aliases'"))
        return issues

    alias_ws = wb["Name Aliases"]
    header_a = alias_ws.cell(row=1, column=1).value
    header_b = alias_ws.cell(row=1, column=2).value
    if (header_a, header_b) != REQUIRED_ALIAS_HEADERS:
        issues.append(Issue(
            "error",
            f"Name Aliases sheet: headers must be {REQUIRED_ALIAS_HEADERS}, got ({header_a!r}, {header_b!r})",
        ))

    canonical_set = set(canonicals)
    alias_raws: list[str] = []
    for r, row in enumerate(alias_ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue
        raw, canon = row[0], row[1] if len(row) > 1 else None
        if not isinstance(raw, str) or not raw.strip():
            issues.append(Issue("error", f"Name Aliases row {r}: blank or non-string raw name"))
            continue
        if not canon or not isinstance(canon, str) or not canon.strip():
            issues.append(Issue("error", f"Name Aliases row {r}: raw {raw!r} points to no canonical"))
            continue
        canon = canon.strip()
        if canon not in canonical_set:
            issues.append(Issue(
                "error",
                f"Name Aliases row {r}: {raw!r} -> {canon!r}, but {canon!r} is not in Employees",
            ))
        alias_raws.append(raw.strip())

    dup_aliases = [n for n, count in Counter(alias_raws).items() if count > 1]
    for dup in dup_aliases:
        issues.append(Issue("warning", f"Duplicate raw name in Name Aliases: {dup!r}"))

    # --- First-name collision warning ---
    by_first: defaultdict[str, list[str]] = defaultdict(list)
    for canonical in canonicals:
        first = canonical.split()[0].lower().rstrip(",")
        by_first[first].append(canonical)
    for first, matches in by_first.items():
        if len(matches) > 1:
            issues.append(Issue(
                "warning",
                f"First-name collision: {matches} — POS entries with just {first!r} "
                f"need an explicit alias or they'll fail to resolve",
            ))

    return issues
