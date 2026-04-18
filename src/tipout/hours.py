from dataclasses import dataclass
from datetime import date
from pathlib import Path
from openpyxl import load_workbook

from tipout.roster import Roster

@dataclass
class HoursEntry:
    canonical: str
    date: date
    hours: float

class MissingHours(RuntimeError): pass
class MissingTips(RuntimeError): pass

def load_hours(path: Path, roster: Roster) -> tuple[list[HoursEntry], list[str]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    entries = []
    unknown = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, d, h = row[0], row[1], row[2]
        if not name:
            continue
        if h is None:
            continue   # skip rows without hours (day off, etc.)
        canon = roster.resolve(name)
        if not canon:
            unknown.append(name)
            continue
        entries.append(HoursEntry(canonical=canon, date=d.date() if hasattr(d, "date") else d, hours=float(h)))
    return entries, sorted(set(unknown))

def validate_join(shift_rows, hours_entries, period_start, period_end):
    """Raise MissingHours/MissingTips if any (canonical, date) in period has one side without the other."""
    from collections import defaultdict
    by_shift = defaultdict(lambda: {"tips": False, "hours": False})
    for r in shift_rows:
        if period_start <= r.date <= period_end:
            by_shift[(r.canonical_name, r.date)]["tips"] = True
    for h in hours_entries:
        if period_start <= h.date <= period_end:
            by_shift[(h.canonical, h.date)]["hours"] = True
    missing_hours = [k for k, v in by_shift.items() if v["tips"] and not v["hours"]]
    missing_tips = [k for k, v in by_shift.items() if v["hours"] and not v["tips"]]
    if missing_hours:
        raise MissingHours(missing_hours)
    if missing_tips:
        raise MissingTips(missing_tips)
