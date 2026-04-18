from dataclasses import dataclass
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

# Canonical header labels. POS sheets vary in case/wording per day
# (e.g. "Net Tip" vs "Net tip", "Serv As" vs "Barback"), so we normalize
# every header we see and alias synonyms to one canonical label.
_HEADER_ALIASES = {
    "cc tips": "CC Tips",
    "party": "Party",
    "cash rcp": "Cash RCP",
    "sa tip out": "SA Tip Out",
    "bar tipout": "Bar Tipout",
    "totaltip out": "TotalTip Out",
    "barback": "Barback",
    "serv as": "Barback",
    "bartender": "Bartender",
    "net tip": "Net tip",
}

EXPECTED_HEADERS = [
    "CC Tips", "SA Tip Out", "Bar Tipout",
    "TotalTip Out", "Barback", "Bartender", "Net tip",
]


class SchemaError(RuntimeError):
    pass

@dataclass
class ShiftRow:
    date: date
    raw_name: str
    cc_tips: float
    party: float
    sa_tip_out: float
    bar_tipout: float
    total_tip_out: float
    barback: float
    bartender: float
    net_tip: float
    is_party: bool  # True if Party column had value (vs Cash RCP)
    canonical_name: str | None = None

def parse_workbook(path: Path) -> list[ShiftRow]:
    wb = load_workbook(path, data_only=True)
    rows: list[ShiftRow] = []
    for sheet_name in wb.sheetnames:
        if sheet_name.strip() in {"Bank Master", "Sheet1"}:
            continue
        ws = wb[sheet_name]
        rows.extend(_parse_sheet(ws))
    return rows

def _parse_sheet(ws) -> list[ShiftRow]:
    blocks = _find_day_blocks(ws)
    out: list[ShiftRow] = []
    for block in blocks:
        out.extend(_parse_block(ws, block))
    return out

def _find_day_blocks(ws) -> list[dict]:
    """Return list of {start_col, header_row, headers, date}.

    Weekly sheets lay out day-blocks in TWO rows of blocks:
    Mon/Tue/Wed at header_row=4 (dates at row 3), then Thu/Fri/Sat/Sun at a
    later header_row (dates one row above). We scan every row for the header
    pattern to catch both groups.
    """
    blocks = []
    max_r = ws.max_row
    max_c = ws.max_column
    for header_row in range(1, max_r + 1):
        c = 1
        while c <= max_c:
            headers = {}
            for offset in range(10):  # a day-block is 10 cols wide
                v = ws.cell(row=header_row, column=c + offset).value
                if isinstance(v, str):
                    canonical = _HEADER_ALIASES.get(v.strip().lower())
                    if canonical is not None:
                        headers.setdefault(canonical, c + offset)
            if all(h in headers for h in EXPECTED_HEADERS):
                date_row = header_row - 1
                date_cell = None
                for offset in range(12):
                    v = ws.cell(row=date_row, column=c + offset).value
                    if hasattr(v, "date"):
                        date_cell = v
                        break
                if date_cell is None:
                    raise ValueError(
                        f"No date found in row {date_row} for day-block starting at "
                        f"col {c} in sheet {ws.title!r}"
                    )
                blocks.append({
                    "start_col": c,
                    "header_row": header_row,
                    "headers": headers,
                    "date": date_cell.date(),
                })
                c = max(headers.values()) + 2
            else:
                c += 1
    if not blocks:
        has_date_anywhere = any(
            hasattr(ws.cell(row=r, column=col).value, "date")
            for r in (3, 42)  # the two date rows we expect
            for col in range(1, max_c + 1)
        )
        if has_date_anywhere:
            raise SchemaError(f"No valid day-blocks in {ws.title!r}")
    return blocks

def _parse_block(ws, block) -> list[ShiftRow]:
    out = []
    # The name column is always the one immediately left of CC Tips.
    # `start_col` varies by block position (Monday: "PM" col, Tuesday/Wed: dayname col)
    # so we can't derive name_col from start_col directly.
    name_col = block["headers"]["CC Tips"] - 1
    # Data rows follow the header row until ~28 rows later (before the next
    # structural section such as the P/O reconciliation block).
    data_start = block["header_row"] + 1
    data_end = block["header_row"] + 29
    for r in range(data_start, data_end):
        raw_name = ws.cell(row=r, column=name_col).value
        if not isinstance(raw_name, str) or not raw_name.strip():
            continue
        def g(label):
            col = block["headers"].get(label)
            if col is None:
                return 0.0
            v = ws.cell(row=r, column=col).value
            return float(v) if isinstance(v, (int, float)) else 0.0
        net = g("Net tip")
        if net == 0.0 and g("CC Tips") == 0.0:
            continue
        name = raw_name.strip()
        party_col_has_value = "Party" in block["headers"] and g("Party") > 0
        name_flags_party = "(party)" in name.lower()
        is_party = party_col_has_value or name_flags_party
        out.append(ShiftRow(
            date=block["date"],
            raw_name=name,
            cc_tips=g("CC Tips"),
            party=g("Party") if party_col_has_value else 0.0,
            sa_tip_out=g("SA Tip Out"),
            bar_tipout=g("Bar Tipout"),
            total_tip_out=g("TotalTip Out"),
            barback=g("Barback"),
            bartender=g("Bartender"),
            net_tip=net,
            is_party=is_party,
        ))
    return out
