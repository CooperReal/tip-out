"""Extractors for bootstrapping tipout inputs from existing hand-kept workbooks."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook


@dataclass
class RosterSnapshot:
    employees: dict[str, str]  # canonical -> role ("" if unknown)
    aliases: dict[str, str]  # raw -> canonical


def extract_roster_from_summary(summary_path: Path) -> RosterSnapshot:
    """Walk the 2-week summary workbook and collect canonical names + aliases.

    Employee rows start at row 5. Col A = canonical. Col B = raw spelling.
    Skip rows where col A is blank.
    """
    wb = load_workbook(summary_path, data_only=True, read_only=True)
    employees: dict[str, str] = {}
    aliases: dict[str, str] = {}

    SKIP_TABS = {"Bank Master", "Sheet1"}

    for sheet_name in wb.sheetnames:
        if sheet_name.strip() in SKIP_TABS:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, max_col=2, values_only=True):
            canonical = row[0]
            raw = row[1] if len(row) > 1 else None
            if not canonical or not isinstance(canonical, str):
                continue
            canonical = canonical.strip()
            if not canonical:
                continue
            employees.setdefault(canonical, "")
            if raw and isinstance(raw, str):
                raw = raw.strip()
                if raw and raw != canonical:
                    aliases[raw] = canonical

    return RosterSnapshot(employees=employees, aliases=aliases)


def extract_roster_from_wvm_daily(wvm_path: Path) -> RosterSnapshot:
    """Seed a roster from the WVM daily worksheet: distinct worker names, role taken
    from the (known) role-group label. No aliases are inferred — the operator dedupes
    genuine misspellings by hand, and must NOT merge same-name people from different
    groups (e.g. 'Carlos' vs 'Carlos Legaspi').
    """
    from tipout.wvm_parser import iter_daily_names

    employees: dict[str, str] = {}
    for name, group in iter_daily_names(wvm_path):
        # First sighting wins; only fill role if we don't already have one.
        if name not in employees or not employees[name]:
            employees[name] = group  # group is '' for junk/blank labels
    return RosterSnapshot(employees=employees, aliases={})


def write_roster(snapshot: RosterSnapshot, out_path: Path) -> None:
    wb = Workbook()
    emp = wb.active
    emp.title = "Employees"
    emp.append(["Canonical Name", "Role", "Active From", "Active To", "Notes"])
    for canonical, role in sorted(snapshot.employees.items()):
        emp.append([canonical, role, None, None, ""])

    alias_ws = wb.create_sheet("Name Aliases")
    alias_ws.append(["Raw Name", "Canonical Name"])
    for raw, canonical in sorted(snapshot.aliases.items()):
        alias_ws.append([raw, canonical])

    wb.save(out_path)
