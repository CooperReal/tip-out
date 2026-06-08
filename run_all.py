"""Drive tipout through every 14-day pay period in the POS file.

For each period:
  - Try to run.
  - If unknown names come back, generate reasonable assumptions,
    update roster.xlsx, retry until no unknowns remain.
  - Record the final grand total vs the hand-reconciled workbook.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Literal

from openpyxl import load_workbook

from tipout.config import Config
from tipout.period import PayPeriod
from tipout.roster import Roster, load_roster
from tipout.runner import UnresolvedNames
from tipout.runner import run as tipout_run
from tipout.summary import _tab_name as period_tab_name

ROOT = Path(__file__).parent
ROSTER = ROOT / "roster.xlsx"
POS = ROOT / "2026 SD Daily Tipout Worksheet.xlsx"
HAND = ROOT / "2026 SD 2 WK Tip Summary By employee.xlsx"
CONFIG = ROOT / "config.yaml"

# Hand workbook has its grand total in col AE (= 31); ours has it in col P (= 16).
HAND_TOTAL_TIPS_COL = 31
TOTAL_TIPS_COL = 16
HAND_TOTAL_ROW_RANGE = range(55, 70)
# Hand workbook has per-employee subtotals; grand total is the one > $1000.
HAND_TOTAL_MIN = 1000
# Each decision can create new canonicals that spawn more ambiguity;
# in practice converges in 1-2 passes.
MAX_UNKNOWN_RETRIES = 5

PERIODS = [
    (date(2025, 12, 29), date(2026, 1, 11)),
    (date(2026, 1, 12), date(2026, 1, 25)),
    (date(2026, 1, 26), date(2026, 2, 8)),
    (date(2026, 2, 9), date(2026, 2, 22)),
    (date(2026, 2, 23), date(2026, 3, 8)),
    (date(2026, 3, 9), date(2026, 3, 22)),
    (date(2026, 3, 23), date(2026, 4, 5)),
    (date(2026, 4, 6), date(2026, 4, 19)),
]


@dataclass
class Decision:
    kind: Literal["alias", "new_employee"]
    canonical: str
    role: str = ""


def guess_mapping(raw: str, roster: Roster) -> Decision:
    """Decide how to resolve an unknown raw name.

    Heuristics:
      - roster resolves it (case-insensitive, after stripping role tags) -> alias
      - first-name matches exactly one canonical -> alias
      - '(Xen)' decoration with no match -> new_employee role 'support'
      - 'party' in the name with no match -> new_employee role 'party'
      - otherwise -> new_employee role 'server'
    """
    stripped = raw.strip()
    low = stripped.lower()

    base = low
    for tag in ("(party)", "(xen)", "(bar back)"):
        base = base.replace(tag, "").strip()
    # 'chuck 03.06.26 party' -> strip trailing date-and-party annotation
    if " party" in base:
        base = base.split(" party", 1)[0].strip()

    resolved = roster.resolve(base)
    if resolved is not None:
        return Decision("alias", resolved)

    first_token = base.split()[0] if base else ""
    if first_token:
        first_matches = [c for c in roster.employees if c.split()[0].lower() == first_token]
        if len(first_matches) == 1:
            return Decision("alias", first_matches[0])

    if "(xen)" in low:
        return Decision("new_employee", stripped, "support")
    if "party" in low:
        return Decision("new_employee", stripped, "party")
    return Decision("new_employee", stripped, "server")


def apply_decisions(decisions: dict[str, Decision]) -> None:
    wb = load_workbook(ROSTER)
    emp_ws = wb["Employees"]
    alias_ws = wb["Name Aliases"]
    existing_emps = {
        row[0].strip() for row in emp_ws.iter_rows(min_row=2, values_only=True) if row and row[0]
    }
    existing_alias_raws = {
        row[0].strip().lower()
        for row in alias_ws.iter_rows(min_row=2, values_only=True)
        if row and row[0]
    }
    for raw, d in decisions.items():
        if d.kind == "new_employee" and d.canonical not in existing_emps:
            emp_ws.append([d.canonical, d.role, None, None, ""])
            existing_emps.add(d.canonical)
        raw_key = raw.strip().lower()
        if raw_key not in existing_alias_raws:
            alias_ws.append([raw, d.canonical])
            existing_alias_raws.add(raw_key)
    wb.save(ROSTER)


def resolve_unknowns(unknowns: list[str]) -> None:
    roster = load_roster(ROSTER)
    decisions: dict[str, Decision] = {}
    for raw in unknowns:
        d = guess_mapping(raw, roster)
        decisions[raw] = d
        if d.kind == "alias":
            print(f"    {raw!r} -> alias to {d.canonical!r}")
        else:
            print(f"    {raw!r} -> NEW employee {d.canonical!r} (role={d.role!r})")
    apply_decisions(decisions)


def run_one_period(cfg: Config, period: PayPeriod) -> str:
    for attempt in range(MAX_UNKNOWN_RETRIES):
        try:
            tipout_run(cfg, POS, period)
            return "ok"
        except UnresolvedNames as exc:
            print(f"  attempt {attempt + 1}: {len(exc.names)} unknown(s)")
            resolve_unknowns(exc.names)
        except ValueError as exc:
            # append_period_tab raises ValueError when the tab already exists.
            return f"skipped: {exc}"
    return f"gave up after {MAX_UNKNOWN_RETRIES} attempts"


def summary_total(summary_path: Path, tab: str) -> float | None:
    """Read the grand total written by summary.py — last row, col AE."""
    if not summary_path.exists():
        return None
    wb = load_workbook(summary_path, data_only=True)
    if tab not in wb.sheetnames:
        return None
    ws = wb[tab]
    v = ws.cell(row=ws.max_row, column=TOTAL_TIPS_COL).value
    return v if isinstance(v, (int, float)) else None


def hand_total(hand_wb, tab: str) -> float | None:
    for candidate in (tab, tab + " ", " " + tab):
        if candidate in hand_wb.sheetnames:
            ws = hand_wb[candidate]
            for r in HAND_TOTAL_ROW_RANGE:
                v = ws.cell(row=r, column=HAND_TOTAL_TIPS_COL).value
                if isinstance(v, (int, float)) and v > HAND_TOTAL_MIN:
                    return v
    return None


def main() -> None:
    cfg = Config.load(CONFIG)
    hand_wb = load_workbook(HAND, data_only=True)

    results = []
    for start, end in PERIODS:
        period = PayPeriod.from_dates(start, end)
        tab = period_tab_name(period)
        print(f"\n==== Period {start} to {end} ({tab}) ====")

        print(f"  {run_one_period(cfg, period)}")

        ours = summary_total(cfg.summary_path, tab)
        hand = hand_total(hand_wb, tab)
        diff = (ours - hand) if (ours is not None and hand is not None) else None
        results.append((tab, hand, ours, diff))

    print("\n\n==== FINAL REPORT ====\n")
    print(f"{'Period':<25} {'Hand':>12} {'Ours':>12} {'Diff':>10}")
    print("-" * 65)
    for tab, hand, ours, diff in results:
        hand_s = f"${hand:,.2f}" if hand is not None else "-"
        ours_s = f"${ours:,.2f}" if ours is not None else "-"
        diff_s = f"${diff:+,.2f}" if diff is not None else "-"
        print(f"{tab:<25} {hand_s:>12} {ours_s:>12} {diff_s:>10}")


if __name__ == "__main__":
    main()
