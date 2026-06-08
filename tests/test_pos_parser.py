from pathlib import Path

import pytest

from tipout.pos_parser import SchemaError, parse_workbook

FIXTURE = Path(__file__).parent / "fixtures" / "tiny_pos.xlsx"


def test_parses_single_day_block():
    rows = parse_workbook(FIXTURE)
    assert len(rows) == 2
    anthony = next(r for r in rows if r.raw_name == "Anthony")
    assert anthony.date.isoformat() == "2025-12-29"
    assert anthony.cc_tips == 583.0
    assert anthony.sa_tip_out == 73.81
    assert anthony.bar_tipout == 34.8
    assert anthony.net_tip == 474.39
    assert anthony.is_party is False


def test_schema_drift_raises(tmp_path):
    from datetime import date as _date

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "test week"
    ws["A1"] = "Surfing Deer"
    ws["B3"] = _date(2026, 1, 5)
    ws["A4"] = "PM"
    ws["B4"] = "Monday"
    ws["C4"] = "CC Tips"
    ws["J4"] = "WRONG HEADER"  # should be "Net tip"
    p = tmp_path / "broken.xlsx"
    wb.save(p)
    with pytest.raises(SchemaError):
        parse_workbook(p)


def test_real_workbook_parses():
    path = Path(__file__).parent / "fixtures" / "real_pos_sample.xlsx"
    if not path.exists():
        pytest.skip("real fixture not present")
    rows = parse_workbook(path)
    # from known data in the first day-block of 12.29:
    dec29 = [r for r in rows if r.date.isoformat() == "2025-12-29"]
    assert any(r.raw_name == "Anthony" and r.net_tip == 474.39 for r in dec29)
    assert any(r.raw_name == "Jake" for r in dec29)
    # Party example from 03.02 for "Patrick (Party)":
    mar2 = [r for r in rows if r.date.isoformat() == "2026-03-02" and "Patrick" in r.raw_name]
    assert any(r.is_party for r in mar2)
