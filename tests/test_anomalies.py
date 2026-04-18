from datetime import date

import pytest
from openpyxl import Workbook, load_workbook

from tipout.anomalies import (
    Anomaly,
    check_all,
    duplicate_rows,
    outlier_rate,
    period_drop,
    server_math,
    tip_pool_imbalance,
    write_report,
)
from tipout.hours import HoursEntry
from tipout.period import PayPeriod
from tipout.pos_parser import ShiftRow
from tipout.roster import load_roster


def _shift(
    d: date,
    raw: str = "x",
    canon: str | None = None,
    cc_tips: float = 0.0,
    party: float = 0.0,
    sa_tip_out: float = 0.0,
    bar_tipout: float = 0.0,
    total_tip_out: float = 0.0,
    barback: float = 0.0,
    bartender: float = 0.0,
    net_tip: float = 0.0,
    is_party: bool = False,
) -> ShiftRow:
    return ShiftRow(
        date=d,
        raw_name=raw,
        cc_tips=cc_tips,
        party=party,
        sa_tip_out=sa_tip_out,
        bar_tipout=bar_tipout,
        total_tip_out=total_tip_out,
        barback=barback,
        bartender=bartender,
        net_tip=net_tip,
        is_party=is_party,
        canonical_name=canon,
    )


# ---------------------------------------------------------------------------
# 1. tip_pool_imbalance
# ---------------------------------------------------------------------------
def test_tip_pool_imbalance_flags_mismatch():
    d = date(2025, 12, 29)
    rows = [
        # server tier: cc_tips > 0; SA distributed = 60 + 40 = 100
        _shift(d, canon="A", cc_tips=500, sa_tip_out=60, total_tip_out=60, net_tip=440),
        _shift(d, canon="B", cc_tips=300, sa_tip_out=40, total_tip_out=40, net_tip=260),
        # support tier: cc_tips == 0, net_tip > 0; received = 80
        _shift(d, canon="C", cc_tips=0, net_tip=80),
    ]
    out = tip_pool_imbalance(rows)
    assert len(out) == 1
    a = out[0]
    assert a.kind == "tip_pool_imbalance"
    assert a.severity == "warn"
    assert a.date == d
    assert a.details["sa_distributed"] == 100.0
    assert a.details["support_received"] == 80.0
    assert a.details["diff"] == 20.0


def test_tip_pool_imbalance_ok_when_balanced():
    d = date(2025, 12, 29)
    rows = [
        _shift(d, canon="A", cc_tips=500, sa_tip_out=50, total_tip_out=50, net_tip=450),
        _shift(d, canon="B", cc_tips=300, sa_tip_out=50, total_tip_out=50, net_tip=250),
        _shift(d, canon="C", cc_tips=0, net_tip=100),
    ]
    assert tip_pool_imbalance(rows) == []


# ---------------------------------------------------------------------------
# 2. server_math
# ---------------------------------------------------------------------------
def test_server_math_flags_bad_math():
    rows = [
        _shift(
            date(2025, 12, 29),
            canon="A",
            cc_tips=100,
            party=0,
            total_tip_out=20,
            net_tip=70,  # expected 80
        ),
    ]
    out = server_math(rows)
    assert len(out) == 1
    assert out[0].kind == "server_math"
    assert out[0].employee == "A"
    assert out[0].details["expected"] == 80.0


def test_server_math_skips_party_zero_tipout():
    rows = [
        _shift(
            date(2025, 12, 29),
            canon="P",
            cc_tips=500,
            party=0,
            total_tip_out=0,
            net_tip=500,
            is_party=True,
        ),
    ]
    assert server_math(rows) == []


# ---------------------------------------------------------------------------
# 3. outlier_rate
# ---------------------------------------------------------------------------
def _period():
    return PayPeriod.from_dates(date(2025, 12, 29), date(2026, 1, 11))


def test_outlier_rate_high():
    p = _period()
    rows = [_shift(date(2025, 12, 29), canon="Anthony Garcia", net_tip=1000)]
    hours = [HoursEntry(canonical="Anthony Garcia", date=date(2025, 12, 29), hours=5.0)]
    out = outlier_rate(rows, hours, p)
    assert len(out) == 1
    assert out[0].employee == "Anthony Garcia"
    assert out[0].details["rate"] == 200.0


def test_outlier_rate_low():
    p = _period()
    rows = [_shift(date(2025, 12, 29), canon="Anthony Garcia", net_tip=20)]
    hours = [HoursEntry(canonical="Anthony Garcia", date=date(2025, 12, 29), hours=5.0)]
    out = outlier_rate(rows, hours, p)
    assert len(out) == 1
    assert out[0].details["rate"] == 4.0


def test_outlier_rate_normal():
    p = _period()
    rows = [_shift(date(2025, 12, 29), canon="Anthony Garcia", net_tip=200)]
    hours = [HoursEntry(canonical="Anthony Garcia", date=date(2025, 12, 29), hours=5.0)]
    assert outlier_rate(rows, hours, p) == []


# ---------------------------------------------------------------------------
# 4. duplicate_rows
# ---------------------------------------------------------------------------
def test_duplicate_rows_flags():
    d = date(2025, 12, 29)
    rows = [
        _shift(d, canon="Anthony Garcia", net_tip=100),
        _shift(d, canon="Anthony Garcia", net_tip=50),
        _shift(d, canon="Jake Purvis", net_tip=80),
    ]
    out = duplicate_rows(rows)
    assert len(out) == 1
    a = out[0]
    assert a.kind == "duplicate_rows"
    assert a.employee == "Anthony Garcia"
    assert a.date == d
    assert a.details["count"] == 2
    assert a.details["net_tips"] == [100.0, 50.0]


# ---------------------------------------------------------------------------
# 5. period_drop
# ---------------------------------------------------------------------------
def test_period_drop_no_history_returns_empty(tiny_roster, tmp_path):
    roster = load_roster(tiny_roster)
    p = _period()
    rows = [_shift(date(2025, 12, 29), canon="Anthony Garcia", net_tip=500)]
    summary_path = tmp_path / "does_not_exist.xlsx"
    assert period_drop(rows, roster, summary_path, p) == []


def test_period_drop_flags_big_drop(tiny_roster, tmp_path):
    """Build a synthetic summary file with 2 prior tabs of ~$3000 totals;
    current period total of $500 should flag (drop > 60%)."""
    roster = load_roster(tiny_roster)
    summary_path = tmp_path / "summary.xlsx"

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Two prior tabs ending before current period start (2025-12-29).
    # Period 1: 12/01 - 12/14
    # Period 2: 12/15 - 12/28
    for tab_name, total in [
        ("12.01 to 12.14.2025", 3000.0),
        ("12.15 to 12.28.2025", 3000.0),
    ]:
        ws = wb.create_sheet(tab_name)
        # Mimic summary layout: rows 1-4 header; row 5+ employee rows;
        # col A canonical, col 31 total tips.
        ws.cell(row=1, column=2, value="Surfing Deer Tip outs")
        ws.cell(row=5, column=1, value="Anthony Garcia")
        ws.cell(row=5, column=2, value="Anthony")
        ws.cell(row=5, column=31, value=total)
    wb.save(summary_path)

    p = PayPeriod.from_dates(date(2025, 12, 29), date(2026, 1, 11))
    rows = [_shift(date(2025, 12, 29), canon="Anthony Garcia", net_tip=500)]

    out = period_drop(rows, roster, summary_path, p)
    assert len(out) == 1
    a = out[0]
    assert a.kind == "period_drop"
    assert a.employee == "Anthony Garcia"
    assert a.details["current"] == 500.0
    assert a.details["rolling_avg"] == 3000.0


# ---------------------------------------------------------------------------
# 6. write_report
# ---------------------------------------------------------------------------
def test_write_report_creates_file(tmp_path):
    p = PayPeriod.from_dates(date(2025, 12, 29), date(2026, 1, 11))
    path = tmp_path / "anomaly_report.xlsx"
    anomalies = [
        Anomaly(
            kind="server_math",
            severity="warn",
            date=date(2025, 12, 29),
            employee="Anthony Garcia",
            message="bad math",
        )
    ]
    write_report(path, anomalies, p)
    assert path.exists()
    wb = load_workbook(path)
    assert wb.sheetnames == ["12.29 to 01.11.2026"]
    ws = wb["12.29 to 01.11.2026"]
    assert ws.cell(row=1, column=1).value == "Kind"
    assert ws.cell(row=1, column=5).value == "Message"
    assert ws.cell(row=2, column=1).value == "server_math"
    assert ws.cell(row=2, column=4).value == "Anthony Garcia"
    assert ws.cell(row=2, column=5).value == "bad math"


def test_write_report_rejects_duplicate_tab(tmp_path):
    p = PayPeriod.from_dates(date(2025, 12, 29), date(2026, 1, 11))
    path = tmp_path / "anomaly_report.xlsx"
    write_report(path, [], p)
    with pytest.raises(ValueError):
        write_report(path, [], p)


# ---------------------------------------------------------------------------
# 7. check_all
# ---------------------------------------------------------------------------
def test_check_all_aggregates_deterministically(tiny_roster, tmp_path):
    """Trigger 3 checks: server_math, outlier_rate, duplicate_rows."""
    roster = load_roster(tiny_roster)
    p = PayPeriod.from_dates(date(2025, 12, 29), date(2026, 1, 11))

    d = date(2025, 12, 29)
    rows = [
        # Bad server math: cc=100 + party=0 - tipout=20 != net=70
        _shift(d, canon="Anthony Garcia", cc_tips=100, total_tip_out=20, net_tip=70),
        # Duplicate (also a good-math row would still flag duplicate)
        _shift(d, canon="Anthony Garcia", cc_tips=80, total_tip_out=0, net_tip=80),
    ]
    # outlier_rate: Jake earns $1000 over 5 hours = $200/hr
    rows.append(_shift(d, canon="Jake Purvis", cc_tips=1000, total_tip_out=0, net_tip=1000))
    hours = [
        HoursEntry(canonical="Anthony Garcia", date=d, hours=8.0),  # 150/8 ≈ 18.75 — in band
        HoursEntry(canonical="Jake Purvis", date=d, hours=5.0),
    ]

    summary_path = tmp_path / "no_summary.xlsx"  # nonexistent — period_drop returns []
    out = check_all(rows, hours, roster, summary_path, p)

    kinds = sorted(a.kind for a in out)
    assert kinds == ["duplicate_rows", "outlier_rate", "server_math"]
    assert len(out) == 3

    # Sort key: (date_or_min, kind, employee_or_'')
    # All three have effective dates: server_math + duplicate_rows on d, outlier_rate has None.
    # outlier_rate (date=None -> date.min) sorts first.
    assert out[0].kind == "outlier_rate"
    assert out[1].kind == "duplicate_rows"
    assert out[2].kind == "server_math"


# ---------------------------------------------------------------------------
# 8. runner integration
# ---------------------------------------------------------------------------
def test_runner_writes_anomaly_report(tiny_runner_env):
    from tipout.config import Config
    from tipout.runner import run

    env = tiny_runner_env
    cfg = Config.load(env["config_path"])
    p = PayPeriod.from_dates(date(2025, 12, 29), date(2026, 1, 11))
    run(cfg, env["pos_path"], env["hours_path"], p)

    anomaly_path = cfg.summary_path.parent / "anomaly_report.xlsx"
    assert anomaly_path.exists()
    wb = load_workbook(anomaly_path)
    assert "12.29 to 01.11.2026" in wb.sheetnames
    ws = wb["12.29 to 01.11.2026"]
    # Header row always present
    assert ws.cell(row=1, column=1).value == "Kind"
