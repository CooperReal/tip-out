from pathlib import Path

from click.testing import CliRunner
from openpyxl import Workbook, load_workbook

from tests.wvm_fixtures import build_wvm_workbook
from tipout.bootstrap import extract_roster_from_summary, extract_roster_from_wvm_daily


def _build_summary(path: Path, tabs: list[tuple[str, list[tuple]]]) -> None:
    """Build a minimal 2-week summary workbook.

    ``tabs`` is a list of (sheet_name, rows) where ``rows`` is a list of
    (canonical, raw) tuples placed starting at row 5.
    """
    wb = Workbook()
    # Drop the default sheet — we'll create our own.
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    for sheet_name, rows in tabs:
        ws = wb.create_sheet(sheet_name)
        # Header rows (1-4) just need to exist; we don't care what's there.
        ws.cell(row=1, column=1, value="2-Week Summary")
        ws.cell(row=4, column=1, value="Canonical")
        ws.cell(row=4, column=2, value="Raw Name")
        for i, (canonical, raw) in enumerate(rows):
            ws.cell(row=5 + i, column=1, value=canonical)
            ws.cell(row=5 + i, column=2, value=raw)
    wb.save(path)


def test_extract_roster_from_minimal_summary(tmp_path):
    path = tmp_path / "summary.xlsx"
    _build_summary(
        path,
        tabs=[
            (
                "12.29 to 01.11.2026",
                [
                    ("Anthony Garcia", "Anthony"),
                    ("Jake Purvis", "Jake"),
                    ("Kristin Bartosic", "Kristin"),
                ],
            ),
            (
                "01.12 to 01.25.2026",
                [
                    # repeats - dedup
                    ("Anthony Garcia", "Anthony"),
                    # same canonical, new raw -> new alias
                    ("Jake Purvis", "Jake P"),
                    # new employee
                    ("Andrew Roberts", "Andrew"),
                ],
            ),
            # Skipped tab
            ("Bank Master", []),
        ],
    )

    snap = extract_roster_from_summary(path)
    assert set(snap.employees.keys()) == {
        "Anthony Garcia",
        "Jake Purvis",
        "Kristin Bartosic",
        "Andrew Roberts",
    }
    # All roles default to ""
    assert all(role == "" for role in snap.employees.values())
    # Aliases include both Jake spellings, but no self-alias
    assert snap.aliases["Anthony"] == "Anthony Garcia"
    assert snap.aliases["Jake"] == "Jake Purvis"
    assert snap.aliases["Jake P"] == "Jake Purvis"
    assert snap.aliases["Kristin"] == "Kristin Bartosic"
    assert snap.aliases["Andrew"] == "Andrew Roberts"
    # No canonical->canonical self-aliases
    for raw, canonical in snap.aliases.items():
        assert raw != canonical


def test_extract_roster_skips_blank_canonical(tmp_path):
    path = tmp_path / "summary.xlsx"
    _build_summary(
        path,
        tabs=[
            (
                "12.29 to 01.11.2026",
                [
                    ("Anthony Garcia", "Anthony"),
                    (None, "GhostName"),         # unresolved row -> skip
                    ("", "AnotherGhost"),        # also skip
                    ("Jake Purvis", "Jake"),
                ],
            )
        ],
    )

    snap = extract_roster_from_summary(path)
    assert set(snap.employees.keys()) == {"Anthony Garcia", "Jake Purvis"}
    assert "GhostName" not in snap.aliases
    assert "AnotherGhost" not in snap.aliases


def test_bootstrap_roster_cli_writes_file(tmp_path):
    from tipout.cli import main

    summary_path = tmp_path / "summary.xlsx"
    _build_summary(
        summary_path,
        tabs=[
            (
                "12.29 to 01.11.2026",
                [
                    ("Anthony Garcia", "Anthony"),
                    ("Jake Purvis", "Jake"),
                ],
            )
        ],
    )

    out_path = tmp_path / "roster.xlsx"
    runner = CliRunner()
    result = runner.invoke(
        main,
        [
            "bootstrap-roster",
            "--from-summary",
            str(summary_path),
            "--out",
            str(out_path),
        ],
    )

    assert result.exit_code == 0, result.output
    assert out_path.exists()
    wb = load_workbook(out_path)
    assert "Employees" in wb.sheetnames
    assert "Name Aliases" in wb.sheetnames

    emp_ws = wb["Employees"]
    assert emp_ws.cell(row=1, column=1).value == "Canonical Name"
    canonicals = [
        emp_ws.cell(row=r, column=1).value
        for r in range(2, emp_ws.max_row + 1)
        if emp_ws.cell(row=r, column=1).value
    ]
    assert set(canonicals) == {"Anthony Garcia", "Jake Purvis"}

    alias_ws = wb["Name Aliases"]
    raws = {
        alias_ws.cell(row=r, column=1).value
        for r in range(2, alias_ws.max_row + 1)
        if alias_ws.cell(row=r, column=1).value
    }
    assert raws == {"Anthony", "Jake"}


def test_bootstrap_roster_refuses_overwrite_without_force(tmp_path):
    from tipout.cli import main

    summary_path = tmp_path / "summary.xlsx"
    _build_summary(
        summary_path,
        tabs=[("12.29 to 01.11.2026", [("Anthony Garcia", "Anthony")])],
    )
    out_path = tmp_path / "roster.xlsx"
    out_path.write_text("pre-existing")

    runner = CliRunner()
    result = runner.invoke(
        main,
        [
            "bootstrap-roster",
            "--from-summary",
            str(summary_path),
            "--out",
            str(out_path),
        ],
    )
    assert result.exit_code != 0
    assert "already exists" in result.output

    # With --force, it succeeds.
    result2 = runner.invoke(
        main,
        [
            "bootstrap-roster",
            "--from-summary",
            str(summary_path),
            "--out",
            str(out_path),
            "--force",
        ],
    )
    assert result2.exit_code == 0, result2.output
    wb = load_workbook(out_path)
    assert "Employees" in wb.sheetnames


def test_extract_from_wvm_daily_collects_names_and_groups(tmp_path):
    p = tmp_path / "wvm.xlsx"
    build_wvm_workbook(p)
    snap = extract_roster_from_wvm_daily(p)
    # distinct names harvested (incl. zero-net "Heather"; excl. P/O strings)
    assert "Ornella" in snap.employees
    assert "Dwayne Graham" in snap.employees
    assert "Cristian Cedeo" in snap.employees
    assert "Total CC Tips" not in snap.employees
    # role taken from known groups; junk '10.19.2222025' must NOT become a role
    assert snap.employees["Ornella"] == "WAIT AM"
    assert snap.employees["Cristian Cedeo"] != "10.19.2222025"
    # harvester seeds no aliases
    assert snap.aliases == {}
