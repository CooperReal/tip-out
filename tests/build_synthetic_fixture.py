"""One-off builder for tests/fixtures/periods/_synthetic/.

Re-run when the synthetic fixture's inputs or expected outputs need to be
regenerated (e.g. after an intentional output-format change). Produces:

  tests/fixtures/periods/_synthetic/
      input_pos.xlsx
      input_hours.xlsx
      input_roster.xlsx
      input_config.yaml
      expected/
          summary.xlsx
          per_employee/<Name>.xlsx

The "expected" outputs are simply whatever the current pipeline produces for
the synthetic inputs. The harness's job is to detect *future drift* against
this snapshot — not to validate current correctness.

Run with: python -m tests.build_synthetic_fixture
"""

from __future__ import annotations

from datetime import date as _date
from pathlib import Path
import shutil

import yaml
from openpyxl import Workbook

from tipout.config import Config
from tipout.period import PayPeriod
from tipout.runner import run as run_pipeline


FIXTURE_DIR = Path(__file__).resolve().parent / "fixtures" / "periods" / "_synthetic"


def _build_pos(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "12.29 to 01.04.2026"
    ws["A1"] = "Surfing Deer Daily Cash"
    ws["B3"] = _date(2025, 12, 29)
    ws["D3"] = "Monday"
    ws["A4"] = "PM"
    ws["B4"] = "Monday"
    ws["C4"] = "CC Tips"
    ws["D4"] = "Party"
    ws["E4"] = "SA Tip Out"
    ws["F4"] = "Bar Tipout"
    ws["G4"] = "TotalTip Out"
    ws["H4"] = "Barback"
    ws["I4"] = "Bartender"
    ws["J4"] = "Net tip"
    # Anthony row
    ws["B5"] = "Anthony"
    ws["C5"] = 583
    ws["E5"] = 73.81
    ws["F5"] = 34.8
    ws["G5"] = 108.61
    ws["J5"] = 474.39
    # Jake row
    ws["B6"] = "Jake"
    ws["C6"] = 219.87
    ws["E6"] = 10.25
    ws["G6"] = 10.25
    ws["I6"] = 65.2
    ws["J6"] = 274.82
    wb.save(path)


def _build_hours(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Hours"
    ws.append(["Employee Name", "Date", "Hours Worked"])
    ws.append(["Anthony Garcia", _date(2025, 12, 29), 7.2])
    ws.append(["Jake Purvis", _date(2025, 12, 29), 6.5])
    wb.save(path)


def _build_roster(path: Path) -> None:
    wb = Workbook()
    emp = wb.active
    emp.title = "Employees"
    emp.append(["Canonical Name", "Role", "Active From", "Active To", "Notes"])
    emp.append(["Anthony Garcia", "server", _date(2025, 1, 1), None, ""])
    emp.append(["Jake Purvis", "bartender", _date(2025, 1, 1), None, ""])
    aliases = wb.create_sheet("Name Aliases")
    aliases.append(["Raw Name", "Canonical Name"])
    aliases.append(["Anthony", "Anthony Garcia"])
    aliases.append(["Jake", "Jake Purvis"])
    wb.save(path)


def _build_config(path: Path) -> None:
    payload = {
        "anchor_date": _date(2025, 12, 29),
        "roster_path": "input_roster.xlsx",
        "summary_path": "output_summary.xlsx",
        "per_employee_dir": "output_per_employee",
        "archive_dir": "output_archive",
    }
    path.write_text(yaml.safe_dump(payload), encoding="utf-8")


def build() -> None:
    if FIXTURE_DIR.exists():
        shutil.rmtree(FIXTURE_DIR)
    FIXTURE_DIR.mkdir(parents=True)

    _build_pos(FIXTURE_DIR / "input_pos.xlsx")
    _build_hours(FIXTURE_DIR / "input_hours.xlsx")
    _build_roster(FIXTURE_DIR / "input_roster.xlsx")
    _build_config(FIXTURE_DIR / "input_config.yaml")

    # Run the pipeline against the just-written inputs to generate the
    # "expected" outputs. We resolve config relative to the fixture dir, so
    # the pipeline writes its outputs INSIDE the fixture dir; we then move
    # those outputs into expected/.
    cfg = Config.load(FIXTURE_DIR / "input_config.yaml")
    period = PayPeriod.from_anchor(_date(2025, 12, 29), _date(2025, 12, 29))
    run_pipeline(
        cfg,
        FIXTURE_DIR / "input_pos.xlsx",
        FIXTURE_DIR / "input_hours.xlsx",
        period,
    )

    # Reorganize: move outputs into expected/.
    expected = FIXTURE_DIR / "expected"
    expected.mkdir(exist_ok=True)
    (expected / "per_employee").mkdir(exist_ok=True)

    shutil.move(str(cfg.summary_path), expected / "summary.xlsx")

    actual_per_emp_dir = cfg.per_employee_dir
    for f in sorted(actual_per_emp_dir.glob("*.xlsx")):
        shutil.move(str(f), expected / "per_employee" / f.name)
    actual_per_emp_dir.rmdir()

    # Anomaly report (optional). Move it into expected/ if the runner produced one.
    anomaly_path = cfg.summary_path.parent / "anomaly_report.xlsx"
    if anomaly_path.exists():
        shutil.move(str(anomaly_path), expected / "anomaly_report.xlsx")

    print(f"Synthetic fixture built at {FIXTURE_DIR}")


if __name__ == "__main__":
    build()
