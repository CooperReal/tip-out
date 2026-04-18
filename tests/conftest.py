from datetime import date as _date
from openpyxl import Workbook
import pytest
import yaml

@pytest.fixture
def tiny_roster(tmp_path):
    wb = Workbook()
    emp = wb.active
    emp.title = "Employees"
    emp.append(["Canonical Name", "Role", "Active From", "Active To", "Notes"])
    emp.append(["Anthony Garcia", "server", _date(2025, 1, 1), None, ""])
    emp.append(["Jake Purvis", "bartender", _date(2025, 1, 1), None, ""])
    emp.append(["Kristin Bartosic", "bartender", _date(2025, 1, 1), None, ""])
    emp.append(["Andrew Roberts", "server", _date(2025, 1, 1), None, ""])
    emp.append(["Andrew Neita", "server", _date(2025, 1, 1), None, ""])
    aliases = wb.create_sheet("Name Aliases")
    aliases.append(["Raw Name", "Canonical Name"])
    aliases.append(["Anthony", "Anthony Garcia"])
    aliases.append(["anthony", "Anthony Garcia"])
    aliases.append(["Jake", "Jake Purvis"])
    aliases.append(["Kristin", "Kristin Bartosic"])
    aliases.append(["kristin", "Kristin Bartosic"])
    path = tmp_path / "roster.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def tiny_hours(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Hours"
    ws.append(["Employee Name", "Date", "Hours Worked"])
    ws.append(["Anthony Garcia", _date(2025, 12, 29), 7.2])
    ws.append(["Jake Purvis", _date(2025, 12, 29), 6.5])
    ws.append(["Kristin Bartosic", _date(2025, 12, 29), 7.0])
    path = tmp_path / "hours.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def tiny_runner_env(tmp_path):
    """Build a self-contained POS + hours + roster + config for end-to-end tests.

    POS and hours are aligned: Anthony + Jake both worked on 12/29 only. No
    Kristin — keeps validate_join clean.
    """
    # --- POS workbook (matches tests/fixtures/tiny_pos.xlsx shape) ---
    pos_wb = Workbook()
    ws = pos_wb.active
    ws.title = "12.29 to 01.04.2026"
    ws["A1"] = "Surfing Deer Daily Cash"
    ws["B3"] = _date(2025, 12, 29)
    ws["D3"] = "Monday"
    ws["A4"] = "PM"
    ws["B4"] = "Monday"
    ws["C4"] = "CC Tips"
    ws["D4"] = "Party"
    ws["E4"] = "SA Tip Out"
    ws["F4"] = "Bar Tipout"
    ws["G4"] = "TotalTip Out"
    ws["H4"] = "Barback"
    ws["I4"] = "Bartender"
    ws["J4"] = "Net tip"
    # Anthony row
    ws["B5"] = "Anthony"
    ws["C5"] = 583
    ws["E5"] = 73.81
    ws["F5"] = 34.8
    ws["G5"] = 108.61
    ws["J5"] = 474.39
    # Jake row
    ws["B6"] = "Jake"
    ws["C6"] = 219.87
    ws["E6"] = 10.25
    ws["G6"] = 10.25
    ws["I6"] = 65.2
    ws["J6"] = 274.82
    pos_path = tmp_path / "pos.xlsx"
    pos_wb.save(pos_path)

    # --- Hours workbook (Anthony + Jake on 12/29 only) ---
    hours_wb = Workbook()
    hws = hours_wb.active
    hws.title = "Hours"
    hws.append(["Employee Name", "Date", "Hours Worked"])
    hws.append(["Anthony Garcia", _date(2025, 12, 29), 7.2])
    hws.append(["Jake Purvis", _date(2025, 12, 29), 6.5])
    hours_path = tmp_path / "hours.xlsx"
    hours_wb.save(hours_path)

    # --- Roster workbook ---
    rwb = Workbook()
    emp = rwb.active
    emp.title = "Employees"
    emp.append(["Canonical Name", "Role", "Active From", "Active To", "Notes"])
    emp.append(["Anthony Garcia", "server", _date(2025, 1, 1), None, ""])
    emp.append(["Jake Purvis", "bartender", _date(2025, 1, 1), None, ""])
    emp.append(["Kristin Bartosic", "bartender", _date(2025, 1, 1), None, ""])
    emp.append(["Andrew Roberts", "server", _date(2025, 1, 1), None, ""])
    emp.append(["Andrew Neita", "server", _date(2025, 1, 1), None, ""])
    aliases = rwb.create_sheet("Name Aliases")
    aliases.append(["Raw Name", "Canonical Name"])
    aliases.append(["Anthony", "Anthony Garcia"])
    aliases.append(["anthony", "Anthony Garcia"])
    aliases.append(["Jake", "Jake Purvis"])
    aliases.append(["Kristin", "Kristin Bartosic"])
    aliases.append(["kristin", "Kristin Bartosic"])
    roster_path = tmp_path / "roster.xlsx"
    rwb.save(roster_path)

    # --- config.yaml ---
    summary_path = tmp_path / "summary.xlsx"
    per_employee_dir = tmp_path / "per-employee"
    archive_dir = tmp_path / "archive"
    config_path = tmp_path / "config.yaml"
    config_path.write_text(
        yaml.safe_dump(
            {
                "anchor_date": _date(2025, 12, 29),
                "roster_path": str(roster_path),
                "summary_path": str(summary_path),
                "per_employee_dir": str(per_employee_dir),
                "archive_dir": str(archive_dir),
            }
        ),
        encoding="utf-8",
    )

    return {
        "config_path": config_path,
        "pos_path": pos_path,
        "hours_path": hours_path,
        "roster_path": roster_path,
        "summary_path": summary_path,
        "per_employee_dir": per_employee_dir,
    }
