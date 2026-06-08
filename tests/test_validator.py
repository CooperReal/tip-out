from datetime import date

from click.testing import CliRunner
from openpyxl import Workbook


def _write_roster(
    path,
    employees,
    aliases,
    *,
    emp_header="Canonical Name",
    alias_headers=("Raw Name", "Canonical Name"),
):
    wb = Workbook()
    emp = wb.active
    emp.title = "Employees"
    emp.append([emp_header, "Role", "Active From", "Active To", "Notes"])
    for canonical in employees:
        emp.append([canonical, "", None, None, ""])
    alias_ws = wb.create_sheet("Name Aliases")
    alias_ws.append(list(alias_headers))
    for raw, canon in aliases:
        alias_ws.append([raw, canon])
    wb.save(path)


def test_clean_roster_has_no_issues(tmp_path):
    from tipout.validator import validate_roster

    path = tmp_path / "roster.xlsx"
    _write_roster(
        path,
        employees=["Anthony Garcia", "Jake Purvis"],
        aliases=[("Anthony", "Anthony Garcia"), ("Jake", "Jake Purvis")],
    )
    assert validate_roster(path) == []


def test_missing_file_is_error(tmp_path):
    from tipout.validator import validate_roster

    issues = validate_roster(tmp_path / "nope.xlsx")
    assert len(issues) == 1
    assert issues[0].severity == "error"
    assert "not found" in issues[0].message


def test_missing_employees_sheet(tmp_path):
    from openpyxl import Workbook
    from tipout.validator import validate_roster

    path = tmp_path / "roster.xlsx"
    wb = Workbook()
    wb.active.title = "Other"
    wb.save(path)

    issues = validate_roster(path)
    assert any(i.severity == "error" and "Employees" in i.message for i in issues)


def test_orphan_alias_is_error(tmp_path):
    from tipout.validator import validate_roster

    path = tmp_path / "roster.xlsx"
    _write_roster(
        path,
        employees=["Anthony Garcia"],
        aliases=[("Anthony", "Anthony Garcia"), ("Jake", "Jake Purvis")],  # Jake not in Employees
    )
    issues = validate_roster(path)
    assert any(
        i.severity == "error" and "Jake Purvis" in i.message and "not in Employees" in i.message
        for i in issues
    )


def test_duplicate_canonical_is_warning(tmp_path):
    from tipout.validator import validate_roster

    path = tmp_path / "roster.xlsx"
    _write_roster(
        path,
        employees=["Anthony Garcia", "Anthony Garcia"],
        aliases=[],
    )
    issues = validate_roster(path)
    assert any(i.severity == "warning" and "Duplicate canonical" in i.message for i in issues)


def test_first_name_collision_is_warning(tmp_path):
    from tipout.validator import validate_roster

    path = tmp_path / "roster.xlsx"
    _write_roster(
        path,
        employees=["Andrew Roberts", "Andrew Neita"],
        aliases=[],
    )
    issues = validate_roster(path)
    assert any(i.severity == "warning" and "First-name collision" in i.message for i in issues)


def test_blank_canonical_is_error(tmp_path):
    from openpyxl import Workbook
    from tipout.validator import validate_roster

    path = tmp_path / "roster.xlsx"
    wb = Workbook()
    emp = wb.active
    emp.title = "Employees"
    emp.append(["Canonical Name", "Role", "Active From", "Active To", "Notes"])
    emp.append(["   ", "", None, None, ""])  # whitespace only
    wb.create_sheet("Name Aliases").append(["Raw Name", "Canonical Name"])
    wb.save(path)

    issues = validate_roster(path)
    assert any(i.severity == "error" and "blank" in i.message for i in issues)


def test_cli_check_roster_clean(tmp_path):
    from tipout.cli import main

    path = tmp_path / "roster.xlsx"
    _write_roster(
        path,
        employees=["Anthony Garcia"],
        aliases=[("Anthony", "Anthony Garcia")],
    )
    result = CliRunner().invoke(main, ["check-roster", str(path)])
    assert result.exit_code == 0
    assert "OK" in result.output


def test_validator_flags_dangling_alias(tmp_path):
    from openpyxl import Workbook

    from tipout.validator import validate_roster

    wb = Workbook()
    emp = wb.active
    emp.title = "Employees"
    emp.append(["Canonical Name", "Role", "Active From", "Active To", "Notes"])
    emp.append(["Real Person", "server", None, None, ""])
    al = wb.create_sheet("Name Aliases")
    al.append(["Raw Name", "Canonical Name"])
    al.append(["Ghosty", "Ghost Person"])  # target not in Employees
    p = tmp_path / "roster.xlsx"
    wb.save(p)
    issues = validate_roster(p)
    assert any(i.severity == "error" and "not in Employees" in i.message for i in issues)


def test_cli_check_roster_reports_errors_and_exits_nonzero(tmp_path):
    from tipout.cli import main

    path = tmp_path / "roster.xlsx"
    _write_roster(
        path,
        employees=["Anthony Garcia"],
        aliases=[("Jake", "Jake Purvis")],  # orphan
    )
    result = CliRunner().invoke(main, ["check-roster", str(path)])
    assert result.exit_code == 1
    assert "ERROR" in result.output
