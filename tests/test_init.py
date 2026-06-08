from pathlib import Path

from click.testing import CliRunner
from openpyxl import Workbook, load_workbook


def _build_summary(path: Path, rows: list[tuple]) -> None:
    """Minimal 2-week summary: one tab, (canonical, raw) rows from row 5."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    ws = wb.create_sheet("12.29 to 01.11.2026")
    ws.cell(row=4, column=1, value="Canonical")
    ws.cell(row=4, column=2, value="Raw Name")
    for i, (canonical, raw) in enumerate(rows):
        ws.cell(row=5 + i, column=1, value=canonical)
        ws.cell(row=5 + i, column=2, value=raw)
    wb.save(path)


def test_init_scaffolds_fresh_project(tmp_path):
    from tipout.cli import main

    runner = CliRunner()
    result = runner.invoke(main, ["init", "--dir", str(tmp_path)])

    assert result.exit_code == 0, result.output
    assert (tmp_path / "config.yaml").exists()
    assert (tmp_path / "roster.xlsx").exists()
    assert (tmp_path / "output").is_dir()

    # config.yaml loads and resolves paths relative to the project dir.
    from tipout.config import Config

    cfg = Config.load(tmp_path / "config.yaml")
    assert cfg.roster_path == (tmp_path / "roster.xlsx").resolve()
    assert cfg.summary_path == (tmp_path / "output" / "summary.xlsx").resolve()

    # Empty roster still has the expected structure.
    wb = load_workbook(tmp_path / "roster.xlsx")
    assert "Employees" in wb.sheetnames
    assert "Name Aliases" in wb.sheetnames
    assert wb["Employees"].cell(row=1, column=1).value == "Canonical Name"


def test_init_seeds_roster_from_summary(tmp_path):
    from tipout.cli import main

    summary_path = tmp_path / "old_summary.xlsx"
    _build_summary(summary_path, [("Anthony Garcia", "Anthony"), ("Jake Purvis", "Jake")])

    project = tmp_path / "proj"
    runner = CliRunner()
    result = runner.invoke(
        main,
        ["init", "--dir", str(project), "--from-summary", str(summary_path)],
    )

    assert result.exit_code == 0, result.output
    wb = load_workbook(project / "roster.xlsx")
    emp_ws = wb["Employees"]
    canonicals = {
        emp_ws.cell(row=r, column=1).value
        for r in range(2, emp_ws.max_row + 1)
        if emp_ws.cell(row=r, column=1).value
    }
    assert canonicals == {"Anthony Garcia", "Jake Purvis"}


def test_init_refuses_overwrite_without_force(tmp_path):
    from tipout.cli import main

    runner = CliRunner()
    first = runner.invoke(main, ["init", "--dir", str(tmp_path)])
    assert first.exit_code == 0, first.output

    second = runner.invoke(main, ["init", "--dir", str(tmp_path)])
    assert second.exit_code != 0
    assert "already exists" in second.output

    forced = runner.invoke(main, ["init", "--dir", str(tmp_path), "--force"])
    assert forced.exit_code == 0, forced.output


def test_init_accepts_custom_anchor(tmp_path):
    from datetime import date

    from tipout.cli import main
    from tipout.config import Config

    runner = CliRunner()
    result = runner.invoke(main, ["init", "--dir", str(tmp_path), "--anchor", "2026-01-12"])
    assert result.exit_code == 0, result.output
    cfg = Config.load(tmp_path / "config.yaml")
    assert cfg.anchor_date == date(2026, 1, 12)
