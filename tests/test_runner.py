from datetime import date
from datetime import date as _date

import pytest
from click.testing import CliRunner
from openpyxl import Workbook, load_workbook


def test_run_end_to_end(tiny_runner_env):
    from tipout.runner import run
    from tipout.config import Config
    from tipout.period import PayPeriod

    env = tiny_runner_env
    cfg = Config.load(env["config_path"])
    period = PayPeriod.from_dates(date(2025, 12, 29), date(2026, 1, 11))

    run(cfg, env["pos_path"], period)

    assert cfg.summary_path.exists()
    wb = load_workbook(cfg.summary_path)
    assert wb.sheetnames == ["12.29 to 01.11.2026"]

    ws = wb["12.29 to 01.11.2026"]
    # Employees live at rows 3..N (after row 1 title + row 2 header). The trailing
    # totals row carries the literal "Daily Total" label — exclude it.
    canonicals = [
        ws.cell(row=r, column=1).value
        for r in range(3, 20)
        if ws.cell(row=r, column=1).value
        and ws.cell(row=r, column=1).value != "Daily Total"
    ]
    assert canonicals == ["Anthony Garcia", "Jake Purvis"]


def test_run_raises_on_unknown_name(tiny_runner_env, tmp_path):
    from tipout.runner import run, UnresolvedNames
    from tipout.config import Config
    from tipout.period import PayPeriod

    env = tiny_runner_env

    broken_roster_path = tmp_path / "broken_roster.xlsx"
    wb = Workbook()
    emp = wb.active
    emp.title = "Employees"
    emp.append(["Canonical Name", "Role", "Active From", "Active To", "Notes"])
    emp.append(["Anthony Garcia", "server", date(2025, 1, 1), None, ""])
    aliases = wb.create_sheet("Name Aliases")
    aliases.append(["Raw Name", "Canonical Name"])
    aliases.append(["Anthony", "Anthony Garcia"])
    wb.save(broken_roster_path)

    cfg = Config.load(env["config_path"])
    cfg.roster_path = broken_roster_path
    period = PayPeriod.from_dates(date(2025, 12, 29), date(2026, 1, 11))

    with pytest.raises(UnresolvedNames) as exc:
        run(cfg, env["pos_path"], period)
    assert "Jake" in exc.value.names


def test_cli_run_writes_unknowns_file_and_exits_nonzero(tiny_runner_env, tmp_path):
    """When a raw name is missing from the roster, the CLI writes unknown_names.txt and exits 1."""
    from tipout.cli import main

    env = tiny_runner_env

    # Replace the roster with one that omits Jake.
    broken_roster = tmp_path / "broken_roster.xlsx"
    wb = Workbook()
    emp = wb.active
    emp.title = "Employees"
    emp.append(["Canonical Name", "Role", "Active From", "Active To", "Notes"])
    emp.append(["Anthony Garcia", "server", date(2025, 1, 1), None, ""])
    aliases = wb.create_sheet("Name Aliases")
    aliases.append(["Raw Name", "Canonical Name"])
    aliases.append(["Anthony", "Anthony Garcia"])
    wb.save(broken_roster)

    # Point config at the broken roster.
    import yaml
    cfg_text = env["config_path"].read_text(encoding="utf-8")
    cfg_data = yaml.safe_load(cfg_text)
    cfg_data["roster_path"] = str(broken_roster)
    env["config_path"].write_text(yaml.safe_dump(cfg_data), encoding="utf-8")

    result = CliRunner().invoke(
        main,
        [
            "run",
            "--period", "2025-12-29:2026-01-11",
            "--config", str(env["config_path"]),
            "--pos", str(env["pos_path"]),
        ],
    )
    assert result.exit_code == 1
    unknowns_path = env["config_path"].parent / "unknown_names.txt"
    assert unknowns_path.exists()
    assert "Jake" in unknowns_path.read_text(encoding="utf-8")


def test_cli_run_success(tiny_runner_env):
    from tipout.cli import main

    env = tiny_runner_env
    result = CliRunner().invoke(
        main,
        [
            "run",
            "--period", "2025-12-29:2026-01-11",
            "--config", str(env["config_path"]),
            "--pos", str(env["pos_path"]),
        ],
    )
    assert result.exit_code == 0, result.output
    assert "Done." in result.output


def test_run_with_hours_populates_per_employee_files(tiny_runner_env):
    from tipout.runner import run
    from tipout.config import Config
    from tipout.period import PayPeriod

    env = tiny_runner_env
    cfg = Config.load(env["config_path"])
    period = PayPeriod.from_dates(date(2025, 12, 29), date(2026, 1, 11))

    run(cfg, env["pos_path"], period, hours_path=env["hours_path"])

    per_emp_dir = cfg.summary_path.parent / "per-employee"
    anthony = load_workbook(per_emp_dir / "Anthony Garcia.xlsx")
    ws = anthony["12.29 to 01.11.2026"]
    # Day 1 = 12/29 is at row 3; Hours Worked is col B.
    assert ws.cell(row=3, column=2).value == 6.0
    # Totals row 17, col B = Hours total
    assert ws.cell(row=17, column=2).value == 6.0
    # $/hr in col J on totals row > 0
    assert ws.cell(row=17, column=10).value > 0


def test_run_without_hours_writes_blank_hours_columns(tiny_runner_env):
    from tipout.runner import run
    from tipout.config import Config
    from tipout.period import PayPeriod

    env = tiny_runner_env
    cfg = Config.load(env["config_path"])
    period = PayPeriod.from_dates(date(2025, 12, 29), date(2026, 1, 11))

    run(cfg, env["pos_path"], period)    # no hours_path

    per_emp_dir = cfg.summary_path.parent / "per-employee"
    anthony = load_workbook(per_emp_dir / "Anthony Garcia.xlsx")
    ws = anthony["12.29 to 01.11.2026"]
    # Header still includes Hours Worked at col B and $/hr at col J — layout is stable.
    assert ws.cell(row=2, column=2).value == "Hours Worked"
    assert ws.cell(row=2, column=10).value == "$/hr"
    # But Hours Worked cells are blank.
    assert ws.cell(row=3, column=2).value is None
    assert ws.cell(row=17, column=2).value is None


def test_run_raises_unresolved_hours_names_before_writing_any_file(
    tiny_runner_env, tmp_path
):
    from tipout.runner import run, UnresolvedHoursNames
    from tipout.config import Config
    from tipout.period import PayPeriod

    env = tiny_runner_env
    # Hours CSV with a name (Stranger Person) that isn't in the roster.
    bad_csv = tmp_path / "bad_hours.csv"
    bad_csv.write_text(
        "STRANGER PERSON - WAIT Mon 12-29-2025 - Sun 01-04-2026,,,,,,,\n"
        "Start Date,Start Time,End Date,End Time,Reported Tips,Regular Hours,Overtime Hours,Duration (Hours)\n"
        '"Mon, 12-29-25",3:00 PM,"Mon, 12-29-25",10:30 PM,0,5.0,0,5.0\n'
        "Total,,,,0,5.0,0,5.0\n",
        encoding="utf-8",
    )

    cfg = Config.load(env["config_path"])
    period = PayPeriod.from_dates(date(2025, 12, 29), date(2026, 1, 11))

    per_emp_dir = cfg.summary_path.parent / "per-employee"
    assert not per_emp_dir.exists()

    with pytest.raises(UnresolvedHoursNames) as exc:
        run(cfg, env["pos_path"], period, hours_path=bad_csv)
    assert "Stranger Person" in exc.value.names
    # Critical: no per-employee files were written.
    assert not per_emp_dir.exists()
    # Same invariant for the summary file: no partial state on disk.
    assert not cfg.summary_path.exists()


def test_run_filters_hours_to_pay_period(tiny_runner_env, tmp_path):
    """Hours rows outside the requested period must not appear in per-employee tabs."""
    from tipout.runner import run
    from tipout.config import Config
    from tipout.period import PayPeriod

    env = tiny_runner_env

    # CSV with a shift INSIDE the period and one OUTSIDE.
    csv_path = tmp_path / "wider_hours.csv"
    csv_path.write_text(
        "ANTHONY GARCIA - WAIT Mon 12-29-2025 - Sun 01-25-2026,,,,,,,\n"
        "Start Date,Start Time,End Date,End Time,Reported Tips,Regular Hours,Overtime Hours,Duration (Hours)\n"
        '"Mon, 12-29-25",3:00 PM,"Mon, 12-29-25",10:30 PM,0,6.0,0,6.0\n'
        '"Mon, 01-12-26",3:00 PM,"Mon, 01-12-26",10:30 PM,0,8.0,0,8.0\n'
        "Total,,,,0,14.0,0,14.0\n",
        encoding="utf-8",
    )

    cfg = Config.load(env["config_path"])
    period = PayPeriod.from_dates(date(2025, 12, 29), date(2026, 1, 11))

    run(cfg, env["pos_path"], period, hours_path=csv_path)

    per_emp_dir = cfg.summary_path.parent / "per-employee"
    anthony = load_workbook(per_emp_dir / "Anthony Garcia.xlsx")
    ws = anthony["12.29 to 01.11.2026"]
    # Hours total = only the 12-29 shift (6.0); the 01-12 shift is outside this period.
    assert ws.cell(row=17, column=2).value == 6.0


def test_cli_run_with_hours_flag(tiny_runner_env):
    from tipout.cli import main

    env = tiny_runner_env
    result = CliRunner().invoke(
        main,
        [
            "run",
            "--period", "2025-12-29:2026-01-11",
            "--config", str(env["config_path"]),
            "--pos", str(env["pos_path"]),
            "--hours", str(env["hours_path"]),
        ],
    )
    assert result.exit_code == 0, result.output
    assert "Done." in result.output

    per_emp_dir = env["summary_path"].parent / "per-employee"
    wb = load_workbook(per_emp_dir / "Anthony Garcia.xlsx")
    ws = wb["12.29 to 01.11.2026"]
    assert ws.cell(row=17, column=2).value == 6.0  # Hours total


def test_wvm_run_writes_summary_only(tiny_wvm_runner_env):
    from tipout.runner import run
    from tipout.config import Config
    from tipout.period import PayPeriod

    env = tiny_wvm_runner_env
    cfg = Config.load(env["config_path"])
    period = PayPeriod.from_dates(date(2025, 12, 29), date(2026, 1, 11))
    run(cfg, env["pos_path"], period, restaurant="wvm")

    wb = load_workbook(env["summary_path"])
    ws = wb["12.29 to 01.11.2026"]
    assert ws["A1"].value.startswith("Watersound Village Market Tip outs")
    # no per-employee directory produced for wvm
    assert not (env["summary_path"].parent / "per-employee").exists()


def test_wvm_run_rejects_hours(tiny_wvm_runner_env):
    from tipout.runner import run
    from tipout.config import Config
    from tipout.period import PayPeriod

    env = tiny_wvm_runner_env
    cfg = Config.load(env["config_path"])
    period = PayPeriod.from_dates(date(2025, 12, 29), date(2026, 1, 11))
    with pytest.raises(ValueError, match="hours"):
        run(cfg, env["pos_path"], period, hours_path=env["pos_path"], restaurant="wvm")


def test_wvm_l56_check_passes_on_clean_env(tiny_wvm_runner_env):
    from tipout.runner import run, DanglingAlias, L56Mismatch
    from tipout.config import Config
    from tipout.period import PayPeriod
    from tipout.roster import load_roster

    env = tiny_wvm_runner_env
    cfg = Config.load(env["config_path"])
    period = PayPeriod.from_dates(date(2025, 12, 29), date(2026, 1, 11))
    run(cfg, env["pos_path"], period, restaurant="wvm")  # no raise
    wb = load_workbook(env["summary_path"])
    ws = wb["12.29 to 01.11.2026"]
    # Daily total for 12.29 (col B = day 1) must equal the tab's 586.56 figure.
    # totals row is the last row; find "Daily Total" label in col A.
    label_rows = [c.row for c in ws["A"] if c.value == "Daily Total"]
    tr = label_rows[0]
    assert round(ws.cell(row=tr, column=2).value, 2) == 586.56  # 162.28 + 424.28


def test_wvm_dangling_alias_raises(tiny_wvm_runner_env):
    from tipout.runner import run, DanglingAlias
    from tipout.config import Config
    from tipout.period import PayPeriod

    env = tiny_wvm_runner_env
    # Point an alias at a canonical absent from Employees -> dangling.
    from openpyxl import load_workbook as _lw
    rwb = _lw(env["roster_path"])
    rwb["Name Aliases"].append(["Ornella", "Ghost Person"])  # overrides Ornella mapping
    rwb.save(env["roster_path"])
    cfg = Config.load(env["config_path"])
    period = PayPeriod.from_dates(date(2025, 12, 29), date(2026, 1, 11))
    with pytest.raises(DanglingAlias):
        run(cfg, env["pos_path"], period, restaurant="wvm")


def test_cli_writes_unknown_hours_file_on_resolution_failure(tiny_runner_env, tmp_path):
    from tipout.cli import main

    env = tiny_runner_env
    bad_csv = tmp_path / "bad_hours.csv"
    bad_csv.write_text(
        "STRANGER PERSON - WAIT Mon 12-29-2025 - Sun 01-04-2026,,,,,,,\n"
        "Start Date,Start Time,End Date,End Time,Reported Tips,Regular Hours,Overtime Hours,Duration (Hours)\n"
        '"Mon, 12-29-25",3:00 PM,"Mon, 12-29-25",10:30 PM,0,5.0,0,5.0\n'
        "Total,,,,0,5.0,0,5.0\n",
        encoding="utf-8",
    )

    result = CliRunner().invoke(
        main,
        [
            "run",
            "--period", "2025-12-29:2026-01-11",
            "--config", str(env["config_path"]),
            "--pos", str(env["pos_path"]),
            "--hours", str(bad_csv),
        ],
    )
    assert result.exit_code == 1
    unknowns_path = env["config_path"].parent / "unknown_hours_names.txt"
    assert unknowns_path.exists()
    assert "Stranger Person" in unknowns_path.read_text(encoding="utf-8")


def test_wvm_l56_mismatch_raises(tiny_wvm_runner_env):
    """Verify L56Mismatch is raised when totals-row Net tip disagrees with worker row sum."""
    from tipout.runner import run, L56Mismatch
    from tipout.config import Config
    from tipout.period import PayPeriod

    env = tiny_wvm_runner_env
    # Load the workbook and corrupt the totals-row Net tip (row 7, column 12)
    wb = load_workbook(env["pos_path"])
    wb["12.29.25"].cell(row=7, column=12, value=999.99)
    wb.save(env["pos_path"])

    cfg = Config.load(env["config_path"])
    period = PayPeriod.from_dates(_date(2025, 12, 29), _date(2026, 1, 11))

    with pytest.raises(L56Mismatch):
        run(cfg, env["pos_path"], period, restaurant="wvm")
