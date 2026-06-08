"""Synthetic WVM daily workbook builder for parser tests.

Mirrors the real `2026 WVM Daily Tip out Worksheet.xlsx` layout (title in A1,
'Date' label in A3 + date in B3, headers in row 4 cols C..L, role-group labels in
col A, names in col B, Net tip in col L, a totals row, then a 'P/O' block) AND its
known breaking quirks (string/typo B3, B3 != tab name, corrupted PM CC header,
negative net, junk col-A group label, trailing-space tab, stray Sheet1, far-column
junk).
"""

from __future__ import annotations

from datetime import date, datetime

from openpyxl import Workbook

HEADERS = {
    3: "AM CC Tips",       # col C
    4: "PM CC TIPS",       # col D
    5: "AM STAFF TIP OUT",
    6: "PM STAFF TIP OUT",
    7: "AM Bar Tipout",
    8: "PM BAR TIPOUT",
    9: "TotalTip Out",
    10: "Serv As",
    11: "Bartender",
    12: "Net tip",         # col L
}


def _new_day(wb, tab_name, b3_value, rows, *, corrupt_pm_cc=False, junk_far_col=False):
    """Add one day-tab. `rows` is a list of (group, name, net) tuples (group '' = carry).

    A totals row (sum of net) and a 'P/O' block are appended automatically.
    """
    ws = wb.create_sheet(tab_name)
    ws["A1"] = "WATERSOUND VILLAGE MARKET"
    ws["A3"] = "Date"
    ws["B3"] = b3_value
    for col, label in HEADERS.items():
        ws.cell(row=4, column=col, value=label)
    if corrupt_pm_cc:
        ws.cell(row=4, column=4, value=25.76)  # PM CC TIPS header overwritten with a number
    r = 5
    total = 0.0
    for group, name, net in rows:
        if group:
            ws.cell(row=r, column=1, value=group)
        ws.cell(row=r, column=2, value=name)
        if net is not None:
            ws.cell(row=r, column=12, value=net)
            total += net
        r += 1
    if junk_far_col:
        ws.cell(row=r - 1, column=16, value=999)  # stray value in far col P (ignored)
    totals_row = r
    ws.cell(row=totals_row, column=12, value=round(total, 2))  # the "L56" figure
    po_row = totals_row + 1
    ws.cell(row=po_row, column=1, value="P/O")
    ws.cell(row=po_row, column=2, value="Total CC Tips")  # non-empty col B past the boundary
    ws.cell(row=po_row, column=12, value=12345)  # must NOT be read as a worker
    return ws


def build_wvm_workbook(path) -> None:
    wb = Workbook()
    del wb["Sheet"]

    # Happy day: a worker in two groups (summed downstream), a zero row (skipped).
    _new_day(wb, "12.29.25", datetime(2025, 12, 29), [
        ("WAIT AM", "Ornella", 162.28),
        ("WAIT AM", "Dwayne Graham", 424.28),
        ("WAIT AM", "Heather", 0),          # zero row -> skipped by parser
        ("BARTNDR", "Dwayne Graham", 50.0), # same person, second group
    ], junk_far_col=True)

    # Stray empty sheet, mid-list (not trailing).
    wb.create_sheet("Sheet1")

    # Trailing-space tab name.
    _new_day(wb, "12.30.25 ", datetime(2025, 12, 30), [
        ("WAIT AM", "Ornella", 100.0),
    ])

    # B3 stored as a typo'd string -> date recovered from tab name.
    _new_day(wb, "01.05.2026", "1/5/226", [
        ("WAIT AM", "Ornella", 80.0),
    ])

    # B3 a real date that DISAGREES with the tab name -> warn, use tab name.
    _new_day(wb, "01.06.2026", datetime(2026, 1, 7), [
        ("WAIT AM", "Ornella", 70.0),
    ])

    # Corrupted PM CC header -> Net tip (col L) still read.
    _new_day(wb, "01.07.2026", datetime(2026, 1, 7), [
        ("WAIT AM", "Ornella", 90.0),
    ], corrupt_pm_cc=True)

    # Negative net tip (correction) -> kept.
    _new_day(wb, "01.08.2026", datetime(2026, 1, 8), [
        ("WAIT AM", "Carlos", -7.83),
    ])

    # Junk col-A group label -> must not become a role for the harvester.
    _new_day(wb, "01.09.2026", datetime(2026, 1, 9), [
        ("F Runner", "Cole Sadler", 30.0),
        ("10.19.2222025", "Cristian Cedeo", 25.0),
    ])

    wb.save(path)
