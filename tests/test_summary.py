from datetime import date, date as _date

import pytest
from openpyxl import Workbook, load_workbook

from tipout.period import PayPeriod
from tipout.pos_parser import ShiftRow
from tipout.roster import Employee, load_roster, Roster
from tipout.summary import (
    DATE_FORMAT,
    TIP_FORMAT,
    _tab_name,
    append_period_tab,
    build_grid,
)


def _shift(d: date, raw: str, canon: str | None, net: float) -> ShiftRow:
    return ShiftRow(
        date=d,
        raw_name=raw,
        cc_tips=0.0,
        party=0.0,
        sa_tip_out=0.0,
        bar_tipout=0.0,
        total_tip_out=0.0,
        barback=0.0,
        bartender=0.0,
        net_tip=net,
        is_party=False,
        canonical_name=canon,
    )


def test_build_grid_basic(tiny_roster):
    """Layout: row 1 title, row 2 header, rows 3..N employees, row N+1 totals.
    Cols: A name, B..O day 1..14, P total."""
    roster = load_roster(tiny_roster)
    period = PayPeriod.from_dates(date(2025, 12, 29), date(2026, 1, 11))
    rows = [
        _shift(date(2025, 12, 29), "Anthony", "Anthony Garcia", 100.00),
        _shift(date(2025, 12, 31), "anthony", "Anthony Garcia", 200.00),
        _shift(date(2025, 12, 30), "Jake", "Jake Purvis", 50.00),
    ]

    grid = build_grid(period, rows, roster)

    # Row 1 title.
    assert grid[0][0] == "Surfing Deer Tip outs 12.29 to 01.11.2026"

    # Row 2 header: Employee | day1 | ... | day14 | Total.
    assert grid[1][0] == "Employee"
    assert grid[1][1] == date(2025, 12, 29)
    assert grid[1][2] == date(2025, 12, 30)
    assert grid[1][14] == date(2026, 1, 11)
    assert grid[1][15] == "Total"

    # 2 header rows + 2 employees with activity (Kristin omitted) + 1 totals row.
    assert len(grid) == 2 + 2 + 1

    # Anthony (grid index 2 = Excel row 3).
    anthony = grid[2]
    assert anthony[0] == "Anthony Garcia"
    assert anthony[1] == 100.00  # day 1
    assert anthony[2] is None  # day 2
    assert anthony[3] == 200.00  # day 3
    assert anthony[15] == 300.00  # total

    # Jake (grid index 3).
    jake = grid[3]
    assert jake[0] == "Jake Purvis"
    assert jake[2] == 50.00
    assert jake[15] == 50.00

    # Totals row.
    totals = grid[-1]
    assert totals[0] == "Daily Total"
    assert totals[1] == 100.00
    assert totals[2] == 50.00
    assert totals[3] == 200.00
    assert totals[15] == 350.00


def test_build_grid_omits_zero_activity(tiny_roster):
    roster = load_roster(tiny_roster)  # 3 employees
    period = PayPeriod.from_dates(date(2025, 12, 29), date(2026, 1, 11))
    rows = [_shift(date(2025, 12, 29), "Anthony", "Anthony Garcia", 100.00)]

    grid = build_grid(period, rows, roster)

    # Only Anthony shows; Jake + Kristin omitted (no activity).
    assert len(grid) == 2 + 1 + 1
    assert grid[2][0] == "Anthony Garcia"


def test_build_grid_respects_roster_order(tiny_roster):
    roster = load_roster(tiny_roster)
    period = PayPeriod.from_dates(date(2025, 12, 29), date(2026, 1, 11))
    rows = [
        _shift(date(2025, 12, 30), "Kristin", "Kristin Bartosic", 75.00),
        _shift(date(2025, 12, 30), "Jake", "Jake Purvis", 50.00),
        _shift(date(2025, 12, 29), "Anthony", "Anthony Garcia", 100.00),
    ]

    grid = build_grid(period, rows, roster)

    assert grid[2][0] == "Anthony Garcia"
    assert grid[3][0] == "Jake Purvis"
    assert grid[4][0] == "Kristin Bartosic"


def test_build_grid_excludes_out_of_period(tiny_roster):
    roster = load_roster(tiny_roster)
    period = PayPeriod.from_dates(date(2025, 12, 29), date(2026, 1, 11))
    rows = [
        _shift(date(2025, 12, 29), "Anthony", "Anthony Garcia", 100.00),
        _shift(date(2026, 1, 12), "Anthony", "Anthony Garcia", 999.00),
    ]

    grid = build_grid(period, rows, roster)

    assert grid[2][0] == "Anthony Garcia"
    assert grid[2][15] == 100.00
    assert grid[-1][15] == 100.00


def test_build_grid_skips_rows_without_canonical(tiny_roster):
    roster = load_roster(tiny_roster)
    period = PayPeriod.from_dates(date(2025, 12, 29), date(2026, 1, 11))
    rows = [
        _shift(date(2025, 12, 29), "Anthony", "Anthony Garcia", 100.00),
        _shift(date(2025, 12, 30), "Mystery Person", None, 42.00),
    ]

    grid = build_grid(period, rows, roster)

    assert grid[2][0] == "Anthony Garcia"
    assert grid[2][15] == 100.00
    assert grid[-1][15] == 100.00


def test_build_grid_rounds_floats(tiny_roster):
    roster = load_roster(tiny_roster)
    period = PayPeriod.from_dates(date(2025, 12, 29), date(2026, 1, 11))
    rows = [
        _shift(date(2025, 12, 29), "Anthony", "Anthony Garcia", 0.1),
        _shift(date(2025, 12, 29), "Anthony", "Anthony Garcia", 0.2),
    ]

    grid = build_grid(period, rows, roster)
    assert grid[2][1] == 0.30
    assert grid[2][15] == 0.30


def test_append_period_tab_creates_new_file(tmp_path, tiny_roster):
    roster = load_roster(tiny_roster)
    period = PayPeriod.from_dates(date(2025, 12, 29), date(2026, 1, 11))
    rows = [
        _shift(date(2025, 12, 29), "Anthony", "Anthony Garcia", 100.00),
        _shift(date(2025, 12, 30), "Jake", "Jake Purvis", 50.00),
    ]
    summary_path = tmp_path / "summary.xlsx"

    append_period_tab(summary_path, period, rows, roster)

    wb = load_workbook(summary_path)
    assert wb.sheetnames == ["12.29 to 01.11.2026"]
    ws = wb["12.29 to 01.11.2026"]

    # Title in A1.
    assert ws.cell(row=1, column=1).value == "Surfing Deer Tip outs 12.29 to 01.11.2026"
    # Anthony lives at row 3 col A; total at row 3 col P (=16).
    assert ws.cell(row=3, column=1).value == "Anthony Garcia"
    assert ws.cell(row=3, column=16).value == 100.00


def test_append_period_tab_applies_formats_and_borders(tmp_path, tiny_roster):
    roster = load_roster(tiny_roster)
    period = PayPeriod.from_dates(date(2025, 12, 29), date(2026, 1, 11))
    rows = [_shift(date(2025, 12, 29), "Anthony", "Anthony Garcia", 100.00)]
    summary_path = tmp_path / "summary.xlsx"

    append_period_tab(summary_path, period, rows, roster)

    wb = load_workbook(summary_path)
    ws = wb["12.29 to 01.11.2026"]

    # Number formats.
    assert ws.cell(row=2, column=2).number_format == DATE_FORMAT  # day header
    assert ws.cell(row=3, column=2).number_format == TIP_FORMAT  # day cell
    assert ws.cell(row=3, column=16).number_format == TIP_FORMAT  # total
    # Frozen pane locks col A + the title/header rows.
    assert ws.freeze_panes == "B3"
    # Tightened col widths.
    assert ws.column_dimensions["A"].width == 26.0
    assert ws.column_dimensions["B"].width == 10.5
    # Every body cell has a thin border on each side.
    body = ws.cell(row=3, column=2).border
    assert body.left.style == "thin"
    assert body.right.style == "thin"
    assert body.top.style == "thin"
    # Day 8 (col I = 9) has a medium left border (week divider).
    day8 = ws.cell(row=3, column=9).border
    assert day8.left.style == "medium"
    # Day 7 (col H = 8) has a medium right border (other side of the divider).
    day7 = ws.cell(row=3, column=8).border
    assert day7.right.style == "medium"


def test_append_period_tab_writes_totals_row(tmp_path, tiny_roster):
    roster = load_roster(tiny_roster)
    period = PayPeriod.from_dates(date(2025, 12, 29), date(2026, 1, 11))
    rows = [
        _shift(date(2025, 12, 29), "Anthony", "Anthony Garcia", 100.00),
        _shift(date(2025, 12, 29), "Jake", "Jake Purvis", 25.00),
        _shift(date(2025, 12, 30), "Kristin", "Kristin Bartosic", 75.00),
    ]
    summary_path = tmp_path / "summary.xlsx"

    append_period_tab(summary_path, period, rows, roster)

    wb = load_workbook(summary_path)
    ws = wb["12.29 to 01.11.2026"]

    # 2 header rows + 3 employees + 1 totals = totals at row 6.
    totals_row = 2 + 3 + 1  # = 6
    assert ws.cell(row=totals_row, column=1).value == "Daily Total"
    assert ws.cell(row=totals_row, column=2).value == 125.00  # day 1
    assert ws.cell(row=totals_row, column=3).value == 75.00  # day 2
    assert ws.cell(row=totals_row, column=16).value == 200.00  # grand total
    assert ws.cell(row=totals_row, column=16).font.bold is True


def test_append_period_tab_preserves_prior(tmp_path, tiny_roster):
    roster = load_roster(tiny_roster)
    summary_path = tmp_path / "summary.xlsx"

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    prior = wb.create_sheet("09.29 to 10.12.2025")
    prior.cell(row=1, column=1, value="PRIOR_SENTINEL")
    wb.save(summary_path)

    period = PayPeriod.from_dates(date(2025, 12, 29), date(2026, 1, 11))
    rows = [_shift(date(2025, 12, 29), "Anthony", "Anthony Garcia", 100.00)]
    append_period_tab(summary_path, period, rows, roster)

    wb2 = load_workbook(summary_path)
    assert "09.29 to 10.12.2025" in wb2.sheetnames
    assert "12.29 to 01.11.2026" in wb2.sheetnames
    assert wb2["09.29 to 10.12.2025"].cell(row=1, column=1).value == "PRIOR_SENTINEL"


def test_append_period_tab_rejects_duplicate(tmp_path, tiny_roster):
    roster = load_roster(tiny_roster)
    summary_path = tmp_path / "summary.xlsx"
    period = PayPeriod.from_dates(date(2025, 12, 29), date(2026, 1, 11))
    rows = [_shift(date(2025, 12, 29), "Anthony", "Anthony Garcia", 100.00)]

    append_period_tab(summary_path, period, rows, roster)
    with pytest.raises(ValueError):
        append_period_tab(summary_path, period, rows, roster)


def test_tab_name_formatting():
    p1 = PayPeriod.from_dates(date(2025, 12, 29), date(2026, 1, 11))
    assert _tab_name(p1) == "12.29 to 01.11.2026"

    p2 = PayPeriod.from_dates(date(2025, 9, 29), date(2025, 10, 12))
    assert _tab_name(p2) == "09.29 to 10.12.2025"


def _one_roster():
    return Roster(employees={"Jane Smith": Employee("Jane Smith", "server")}, aliases={})


def _one_row():
    r = ShiftRow(
        date=_date(2025, 12, 29),
        raw_name="Jane",
        cc_tips=0.0,
        party=0.0,
        sa_tip_out=0.0,
        bar_tipout=0.0,
        total_tip_out=0.0,
        barback=0.0,
        bartender=0.0,
        net_tip=100.0,
        is_party=False,
    )
    r.canonical_name = "Jane Smith"
    return r


def test_build_grid_title_uses_restaurant_name():
    period = PayPeriod.from_dates(_date(2025, 12, 29), _date(2026, 1, 11))
    grid = build_grid(
        period, [_one_row()], _one_roster(), restaurant_name="Watersound Village Market"
    )
    assert grid[0][0].startswith("Watersound Village Market Tip outs")


def test_build_grid_title_defaults_to_surfing_deer():
    period = PayPeriod.from_dates(_date(2025, 12, 29), _date(2026, 1, 11))
    grid = build_grid(period, [_one_row()], _one_roster())
    assert grid[0][0].startswith("Surfing Deer Tip outs")
