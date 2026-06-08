from datetime import date

import pytest
from openpyxl import Workbook

from tests.wvm_fixtures import build_wvm_workbook
from tipout.wvm_parser import WvmFormatError, iter_daily_names, parse_workbook, read_day_net_totals


@pytest.fixture
def wvm_path(tmp_path):
    p = tmp_path / "wvm.xlsx"
    build_wvm_workbook(p)
    return p


def test_fixture_builds_expected_sheets(wvm_path):
    from openpyxl import load_workbook
    names = [n.strip() for n in load_workbook(wvm_path).sheetnames]
    assert "12.29.25" in names
    assert "Sheet1" in names
    assert "01.08.2026" in names


def test_parse_reads_net_tip_and_date_from_happy_tab(wvm_path):
    rows = parse_workbook(wvm_path)
    by = [(r.date, r.raw_name, r.net_tip) for r in rows]
    assert (date(2025, 12, 29), "Ornella", 162.28) in by
    assert (date(2025, 12, 29), "Dwayne Graham", 424.28) in by


def test_parse_skips_zero_rows(wvm_path):
    rows = parse_workbook(wvm_path)
    assert not [r for r in rows if r.raw_name == "Heather"]  # net 0 -> skipped


def test_parse_skips_sheet1(wvm_path):
    rows = parse_workbook(wvm_path)
    # Sheet1 has no date-shaped name; nothing should come from it (no crash).
    assert all(r.date is not None for r in rows)


def test_parse_does_not_read_po_block(wvm_path):
    rows = parse_workbook(wvm_path)
    assert not [r for r in rows if r.raw_name == "Total CC Tips"]


def test_parse_keeps_only_net_tip_other_fields_zero(wvm_path):
    rows = parse_workbook(wvm_path)
    r = next(r for r in rows if r.raw_name == "Ornella" and r.date == date(2025, 12, 29))
    assert r.cc_tips == 0.0 and r.bar_tipout == 0.0 and r.is_party is False


def test_string_b3_recovered_from_tab_name(wvm_path, capsys):
    rows = parse_workbook(wvm_path)
    assert any(r.date == date(2026, 1, 5) and r.raw_name == "Ornella" for r in rows)
    assert "is not a date" in capsys.readouterr().err


def test_b3_mismatch_uses_tab_name_and_warns(wvm_path, capsys):
    rows = parse_workbook(wvm_path)
    # tab 01.06.2026 has B3 = 2026-01-07; tab name must win.
    assert any(r.date == date(2026, 1, 6) and r.raw_name == "Ornella" for r in rows)
    assert "!= tab-name date" in capsys.readouterr().err


def test_corrupted_pm_cc_header_still_reads_net(wvm_path):
    rows = parse_workbook(wvm_path)
    assert any(r.date == date(2026, 1, 7) and r.net_tip == 90.0 for r in rows)


def test_negative_net_tip_is_kept(wvm_path):
    rows = parse_workbook(wvm_path)
    assert any(r.raw_name == "Carlos" and r.net_tip == -7.83 for r in rows)


def test_trailing_space_tab_parsed(wvm_path):
    rows = parse_workbook(wvm_path)
    assert any(r.date == date(2025, 12, 30) for r in rows)


def test_same_person_two_groups_yields_two_rows(wvm_path):
    rows = parse_workbook(wvm_path)
    dwayne = [r for r in rows if r.raw_name == "Dwayne Graham" and r.date == date(2025, 12, 29)]
    assert len(dwayne) == 2
    assert sorted(r.net_tip for r in dwayne) == [50.0, 424.28]


def test_read_day_net_totals_returns_sheet_totals(wvm_path):
    totals = read_day_net_totals(wvm_path)
    # 12.29.25 totals-row Net tip = 162.28 + 424.28 + 0 + 50.0 = 636.56
    assert round(totals[date(2025, 12, 29)], 2) == 636.56


def test_parse_non_wvm_file_raises(tmp_path):
    wb = Workbook()
    wb.active.title = "NotADate"
    wb.active["A1"] = "something else"
    p = tmp_path / "notwvm.xlsx"
    wb.save(p)
    with pytest.raises(WvmFormatError):
        parse_workbook(p)


def test_iter_daily_names_includes_zero_rows_and_groups(wvm_path):
    pairs = list(iter_daily_names(wvm_path))
    names = {n for n, _g in pairs}
    assert "Heather" in names          # zero-net row still yielded (unlike parse_workbook)
    assert "Cristian Cedeo" in names
    # junk col-A label must never surface as a group
    assert all(g != "10.19.2222025" for _n, g in pairs)
    assert ("Ornella", "WAIT AM") in pairs
