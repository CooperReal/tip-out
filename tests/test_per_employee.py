from datetime import date

import pytest
from openpyxl import load_workbook

from tipout.per_employee import (
    DATE_FORMAT,
    HOURS_FORMAT,
    TIP_FORMAT,
    _safe_filename,
    append_period_tab_for_employee,
    build_grid,
)
from tipout.period import PayPeriod
from tipout.pos_parser import ShiftRow


def _shift(d: date, canon: str | None, **vals) -> ShiftRow:
    base = dict(
        cc_tips=0.0, party=0.0, sa_tip_out=0.0, bar_tipout=0.0,
        total_tip_out=0.0, barback=0.0, bartender=0.0, net_tip=0.0,
        is_party=False,
    )
    base.update(vals)
    return ShiftRow(
        date=d,
        raw_name=canon or "raw",
        canonical_name=canon,
        **base,
    )


def test_build_grid_layout_and_totals():
    period = PayPeriod.from_dates(date(2025, 12, 29), date(2026, 1, 11))
    rows = [
        _shift(
            date(2025, 12, 30), "Yvonne Lewis",
            cc_tips=231.0, sa_tip_out=10.53, bar_tipout=33.90,
            total_tip_out=44.43, bartender=105.80, net_tip=292.37,
        ),
        _shift(
            date(2025, 12, 31), "Yvonne Lewis",
            cc_tips=280.57, sa_tip_out=13.38, total_tip_out=13.38,
            bartender=72.82, net_tip=340.01,
        ),
    ]

    grid = build_grid(period, "Yvonne Lewis", rows)

    # Title in R1C1.
    assert "Yvonne Lewis" in grid[0][0]
    assert "12.29 to 01.11.2026" in grid[0][0]

    # Header row.
    assert grid[1] == [
        "Date", "Hours Worked", "CC Tips", "SA Tip Out", "Bar Tipout",
        "Total Tip Out", "Barback", "Bartender", "Net Tip", "$/hr",
    ]

    # 2 header rows + 14 day rows + 1 totals row = 17 rows.
    assert len(grid) == 17

    # Day 1 = 12/29 = no shift -> all None except date.
    day1 = grid[2]
    assert day1[0] == date(2025, 12, 29)
    assert all(v is None for v in day1[1:])

    # Day 2 = 12/30 = first shift. CC Tips is now col index 2 (B is Hours).
    day2 = grid[3]
    assert day2[0] == date(2025, 12, 30)
    assert day2[1] is None       # Hours Worked (not provided)
    assert day2[2] == 231.0      # CC Tips
    assert day2[3] == 10.53      # SA Tip Out
    assert day2[4] == 33.90      # Bar Tipout
    assert day2[5] == 44.43      # Total Tip Out
    assert day2[6] is None       # Barback (zero)
    assert day2[7] == 105.80     # Bartender
    assert day2[8] == 292.37     # Net Tip
    assert day2[9] is None       # $/hr (day rows always blank)

    # Totals row.
    totals = grid[-1]
    assert totals[0] == "Total"
    assert totals[1] is None     # Hours total — none provided
    assert totals[2] == 511.57   # CC Tips total
    assert totals[8] == 632.38   # Net Tip total
    assert totals[9] is None     # $/hr blank when hours total is 0


def test_build_grid_omits_other_employees():
    """Shifts for different canonicals must be ignored."""
    period = PayPeriod.from_dates(date(2025, 12, 29), date(2026, 1, 11))
    rows = [
        _shift(date(2025, 12, 30), "Yvonne Lewis", net_tip=100.0),
        _shift(date(2025, 12, 30), "Other Person", net_tip=999.0),
    ]

    grid = build_grid(period, "Yvonne Lewis", rows)
    assert grid[-1][8] == 100.0   # totals: only Yvonne counted


def test_build_grid_excludes_out_of_period():
    period = PayPeriod.from_dates(date(2025, 12, 29), date(2026, 1, 11))
    rows = [
        _shift(date(2025, 12, 29), "Yvonne Lewis", net_tip=50.0),
        _shift(date(2026, 1, 12), "Yvonne Lewis", net_tip=999.0),  # one day past end
    ]

    grid = build_grid(period, "Yvonne Lewis", rows)
    assert grid[-1][8] == 50.0


def test_append_period_tab_writes_file_with_styling(tmp_path):
    period = PayPeriod.from_dates(date(2025, 12, 29), date(2026, 1, 11))
    rows = [
        _shift(
            date(2025, 12, 30), "Yvonne Lewis",
            cc_tips=231.0, net_tip=292.37,
        ),
    ]

    written = append_period_tab_for_employee(tmp_path, period, "Yvonne Lewis", rows)

    assert written == tmp_path / "per-employee" / "Yvonne Lewis.xlsx"
    assert written.exists()

    wb = load_workbook(written)
    assert wb.sheetnames == ["12.29 to 01.11.2026"]
    ws = wb["12.29 to 01.11.2026"]

    # Title.
    assert "Yvonne Lewis" in ws.cell(row=1, column=1).value

    # Day 2 = 12/30 has CC Tips populated. (openpyxl reads back date → datetime.)
    cell_date = ws.cell(row=4, column=1).value
    assert (cell_date.year, cell_date.month, cell_date.day) == (2025, 12, 30)
    # CC Tips now at col 3 (col 2 = Hours Worked, blank here).
    assert ws.cell(row=4, column=2).value is None
    assert ws.cell(row=4, column=3).value == 231.0

    # Number formats applied.
    assert ws.cell(row=3, column=1).number_format == DATE_FORMAT
    assert ws.cell(row=4, column=3).number_format == TIP_FORMAT
    # Hours Worked (col B) is a plain number, not currency — day rows and total.
    assert ws.cell(row=4, column=2).number_format == HOURS_FORMAT
    assert ws.cell(row=17, column=2).number_format == HOURS_FORMAT

    # Frozen pane locks title + header.
    assert ws.freeze_panes == "A3"

    # Borders on body cells.
    assert ws.cell(row=4, column=2).border.left.style == "thin"

    # Totals row at row 17, bold.
    assert ws.cell(row=17, column=1).value == "Total"
    assert ws.cell(row=17, column=3).font.bold is True


def test_append_period_tab_appends_without_overwriting(tmp_path):
    period1 = PayPeriod.from_dates(date(2025, 12, 29), date(2026, 1, 11))
    period2 = PayPeriod.from_dates(date(2026, 1, 12), date(2026, 1, 25))
    rows = [
        _shift(date(2025, 12, 29), "Yvonne Lewis", net_tip=50.0),
        _shift(date(2026, 1, 12), "Yvonne Lewis", net_tip=80.0),
    ]

    append_period_tab_for_employee(tmp_path, period1, "Yvonne Lewis", rows)
    append_period_tab_for_employee(tmp_path, period2, "Yvonne Lewis", rows)

    wb = load_workbook(tmp_path / "per-employee" / "Yvonne Lewis.xlsx")
    assert wb.sheetnames == ["12.29 to 01.11.2026", "01.12 to 01.25.2026"]


def test_append_period_tab_rejects_duplicate_period(tmp_path):
    period = PayPeriod.from_dates(date(2025, 12, 29), date(2026, 1, 11))
    rows = [_shift(date(2025, 12, 29), "Yvonne Lewis", net_tip=50.0)]

    append_period_tab_for_employee(tmp_path, period, "Yvonne Lewis", rows)
    with pytest.raises(ValueError):
        append_period_tab_for_employee(tmp_path, period, "Yvonne Lewis", rows)


def test_safe_filename_strips_illegal_chars():
    assert _safe_filename("Anthony Garcia") == "Anthony Garcia"
    assert _safe_filename('Bad/Name?') == "Bad_Name_"
    assert _safe_filename("Marcus, Eric") == "Marcus, Eric"  # comma is legal


def test_build_grid_with_hours_populates_b_and_j():
    period = PayPeriod.from_dates(date(2025, 12, 29), date(2026, 1, 11))
    rows = [
        _shift(date(2025, 12, 30), "Yvonne Lewis", net_tip=200.0),
        _shift(date(2025, 12, 31), "Yvonne Lewis", net_tip=300.0),
    ]
    hours_by_date = {
        date(2025, 12, 30): 7.5,
        date(2025, 12, 31): 5.0,
        # Day with hours but no tips — still surfaces in Hours column.
        date(2026, 1, 1): 4.0,
    }

    grid = build_grid(period, "Yvonne Lewis", rows, hours_by_date=hours_by_date)

    # Day 2 = 12/30
    assert grid[3][1] == 7.5
    # Day 3 = 12/31
    assert grid[4][1] == 5.0
    # Day 4 = 1/1 (no tip shift, but hours present)
    assert grid[5][1] == 4.0
    # Day 5 onward: no hours, no tips → blank
    assert grid[6][1] is None

    # Totals row.
    totals = grid[-1]
    assert totals[1] == 16.5    # hours total
    # $/hr = Net Tip total (500.0) / Hours total (16.5) → 30.30
    assert totals[9] == round(500.0 / 16.5, 2)


def test_build_grid_with_zero_hours_total_leaves_per_hr_blank():
    period = PayPeriod.from_dates(date(2025, 12, 29), date(2026, 1, 11))
    rows = [_shift(date(2025, 12, 30), "Yvonne Lewis", net_tip=50.0)]
    grid = build_grid(period, "Yvonne Lewis", rows, hours_by_date={})
    totals = grid[-1]
    assert totals[1] is None    # no hours
    assert totals[9] is None    # $/hr left blank, not zero


def test_hours_column_uses_plain_number_not_currency_format(tmp_path):
    """Hours Worked (col B) renders as a plain number; tips and $/hr stay currency.

    Regression for the change request: col B was inheriting the currency
    format, so e.g. 7.55 hours displayed as $7.55. Mirrors Yvonne's example.
    """
    period = PayPeriod.from_dates(date(2025, 12, 29), date(2026, 1, 11))
    rows = [_shift(date(2025, 12, 30), "Yvonne Lewis", net_tip=292.37)]
    hours_by_date = {date(2025, 12, 30): 7.55, date(2025, 12, 31): 6.6}

    written = append_period_tab_for_employee(
        tmp_path, period, "Yvonne Lewis", rows, hours_by_date=hours_by_date
    )
    ws = load_workbook(written)["12.29 to 01.11.2026"]

    # Hours column (B) is a plain number on day rows AND the totals row.
    for r in range(3, 18):  # rows 3..17 = 14 day rows + totals
        assert ws.cell(row=r, column=2).number_format == HOURS_FORMAT
    assert ws.cell(row=4, column=2).value == 7.55       # value unchanged
    assert ws.cell(row=17, column=2).value == 14.15     # hours total

    # Tip columns (C..I) keep currency on day rows and the totals row.
    for c in range(3, 10):
        assert ws.cell(row=4, column=c).number_format == TIP_FORMAT
        assert ws.cell(row=17, column=c).number_format == TIP_FORMAT

    # $/hr (col J) on the totals row stays currency.
    assert ws.cell(row=17, column=10).number_format == TIP_FORMAT
