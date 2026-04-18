import json
from datetime import date

import pytest
from click.testing import CliRunner
from openpyxl import Workbook, load_workbook


def _broken_roster(path):
    """Overwrite the roster at `path` so it omits Jake (forces an unresolved name)."""
    wb = Workbook()
    emp = wb.active
    emp.title = "Employees"
    emp.append(["Canonical Name", "Role", "Active From", "Active To", "Notes"])
    emp.append(["Anthony Garcia", "server", date(2025, 1, 1), None, ""])
    aliases = wb.create_sheet("Name Aliases")
    aliases.append(["Raw Name", "Canonical Name"])
    aliases.append(["Anthony", "Anthony Garcia"])
    wb.save(path)


def _invoke_run(env, json_flag=True):
    from tipout.cli import main
    args = [
        "run",
        "--period", "2025-12-29:2026-01-11",
        "--config", str(env["config_path"]),
        "--pos", str(env["pos_path"]),
        "--hours", str(env["hours_path"]),
    ]
    if json_flag:
        args.append("--json")
    return CliRunner().invoke(main, args, catch_exceptions=False)


def test_run_json_unknown_writes_pending_and_nonzero(tiny_runner_env):
    env = tiny_runner_env
    _broken_roster(env["roster_path"])

    result = _invoke_run(env, json_flag=True)

    assert result.exit_code == 1, result.output
    payload = json.loads(result.output)
    assert payload["status"] == "awaiting_input"
    assert payload["reason"] == "unresolved_names"
    assert "Jake" in payload["unresolved_names"]

    pending = env["config_path"].parent / "pending_names.json"
    assert pending.exists()
    written = json.loads(pending.read_text())
    assert written["status"] == "awaiting_input"
    assert "Jake" in written["unresolved_names"]

    # No deliverables emitted
    assert not env["summary_path"].exists()
    assert not env["per_employee_dir"].exists() or not any(env["per_employee_dir"].iterdir())


def test_run_json_success(tiny_runner_env):
    env = tiny_runner_env
    result = _invoke_run(env, json_flag=True)

    assert result.exit_code == 0, result.output
    payload = json.loads(result.output)
    assert payload["status"] == "success"
    assert payload["period"] == "2025-12-29:2026-01-11"
    assert env["summary_path"].exists()


def test_resolve_pending_new_employee(tiny_runner_env):
    from tipout.cli import main

    env = tiny_runner_env
    _broken_roster(env["roster_path"])

    # Trigger pending_names.json
    result = _invoke_run(env, json_flag=True)
    assert result.exit_code == 1

    pending_names = env["config_path"].parent / "pending_names.json"
    assert pending_names.exists()

    # Operator answers Jake -> new employee
    answers_path = env["config_path"].parent / "pending_answers.json"
    answers_path.write_text(json.dumps({
        "Jake": {
            "decision": "new_employee",
            "canonical_name": "Jake Purvis",
            "role": "bartender",
        }
    }))

    result = CliRunner().invoke(
        main, ["resolve-pending", "--config", str(env["config_path"])],
        catch_exceptions=False,
    )
    assert result.exit_code == 0, result.output

    # Roster has Jake and the alias
    wb = load_workbook(env["roster_path"])
    emp_names = [
        wb["Employees"].cell(row=r, column=1).value
        for r in range(2, wb["Employees"].max_row + 1)
    ]
    assert "Jake Purvis" in emp_names
    alias_pairs = [
        (wb["Name Aliases"].cell(row=r, column=1).value,
         wb["Name Aliases"].cell(row=r, column=2).value)
        for r in range(2, wb["Name Aliases"].max_row + 1)
    ]
    assert ("Jake", "Jake Purvis") in alias_pairs

    # Pending files cleaned up
    assert not pending_names.exists()
    assert not answers_path.exists()


def test_resolve_pending_ignore(tiny_runner_env):
    from tipout.cli import main
    from tipout.runner import IGNORE_SENTINEL

    env = tiny_runner_env
    _broken_roster(env["roster_path"])

    result = _invoke_run(env, json_flag=True)
    assert result.exit_code == 1

    answers_path = env["config_path"].parent / "pending_answers.json"
    answers_path.write_text(json.dumps({
        "Jake": {"decision": "ignore"}
    }))

    result = CliRunner().invoke(
        main, ["resolve-pending", "--config", str(env["config_path"])],
        catch_exceptions=False,
    )
    assert result.exit_code == 0, result.output

    wb = load_workbook(env["roster_path"])
    alias_pairs = [
        (wb["Name Aliases"].cell(row=r, column=1).value,
         wb["Name Aliases"].cell(row=r, column=2).value)
        for r in range(2, wb["Name Aliases"].max_row + 1)
    ]
    assert ("Jake", IGNORE_SENTINEL) in alias_pairs

    # Re-run pipeline - should now succeed (Jake's POS row filtered out;
    # hours file has Jake but the broken roster also lacks Jake from hours
    # mapping, so we need to also drop Jake from hours. The hours file maps
    # 'Jake Purvis' which is still unresolved in the broken roster.) For the
    # purpose of this test, also remove Jake from hours so the pipeline runs.
    hwb = Workbook()
    hws = hwb.active
    hws.title = "Hours"
    hws.append(["Employee Name", "Date", "Hours Worked"])
    hws.append(["Anthony Garcia", date(2025, 12, 29), 7.2])
    hwb.save(env["hours_path"])

    result = _invoke_run(env, json_flag=True)
    assert result.exit_code == 0, result.output
    payload = json.loads(result.output)
    assert payload["status"] == "success"


def test_ignore_filters_pos_only_name(tiny_runner_env):
    """A POS name mapped to IGNORE_SENTINEL + no hours row should run cleanly."""
    from tipout.runner import IGNORE_SENTINEL

    env = tiny_runner_env

    # Roster: Anthony as full employee; Jake only as an alias -> IGNORE_SENTINEL.
    wb = Workbook()
    emp = wb.active
    emp.title = "Employees"
    emp.append(["Canonical Name", "Role", "Active From", "Active To", "Notes"])
    emp.append(["Anthony Garcia", "server", date(2025, 1, 1), None, ""])
    aliases = wb.create_sheet("Name Aliases")
    aliases.append(["Raw Name", "Canonical Name"])
    aliases.append(["Anthony", "Anthony Garcia"])
    aliases.append(["Jake", IGNORE_SENTINEL])
    wb.save(env["roster_path"])

    # Hours: only Anthony -- no Jake row at all.
    hwb = Workbook()
    hws = hwb.active
    hws.title = "Hours"
    hws.append(["Employee Name", "Date", "Hours Worked"])
    hws.append(["Anthony Garcia", date(2025, 12, 29), 7.2])
    hwb.save(env["hours_path"])

    result = _invoke_run(env, json_flag=True)
    assert result.exit_code == 0, result.output
    payload = json.loads(result.output)
    assert payload["status"] == "success"

    # Summary has Anthony but not Jake.
    assert env["summary_path"].exists()
    swb = load_workbook(env["summary_path"])
    ws = swb[swb.sheetnames[0]]
    canonicals = [
        ws.cell(row=r, column=1).value
        for r in range(5, 25)
        if ws.cell(row=r, column=1).value
    ]
    assert "Anthony Garcia" in canonicals
    assert "Jake Purvis" not in canonicals
    assert IGNORE_SENTINEL not in canonicals


def test_resolve_pending_alias_requires_existing_canonical(tiny_runner_env):
    from tipout.cli import main

    env = tiny_runner_env
    _broken_roster(env["roster_path"])

    result = _invoke_run(env, json_flag=True)
    assert result.exit_code == 1

    answers_path = env["config_path"].parent / "pending_answers.json"
    answers_path.write_text(json.dumps({
        "Jake": {"decision": "alias", "canonical_name": "Nobody Here"}
    }))

    result = CliRunner().invoke(
        main, ["resolve-pending", "--config", str(env["config_path"])],
        catch_exceptions=False,
    )
    assert result.exit_code != 0
    assert "Nobody Here" in result.output or "Nobody Here" in (result.stderr_bytes or b"").decode()
